#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Smart Chunking Module for Policy Extraction

This module provides advanced document chunking strategies specifically designed
for policy extraction from large documents. It includes:

- Policy structure analysis from table of contents and initial pages
- Image-based analysis of initial pages for policy identification
- Smart document segmentation based on policy identification
- Automated blank and title page removal
- Efficient policy boundary detection
"""

import re
import os
import io
import base64
import logging
import shutil  # Add import for directory removal
from typing import List, Dict, Any, Optional, Tuple, Set, Union
import fitz  # PyMuPDF
import json
import tempfile
from dataclasses import dataclass, field
from collections import defaultdict
from pydantic import BaseModel, Field
from PIL import Image
from openai import AsyncAzureOpenAI
import asyncio
from io import BytesIO
from dotenv import load_dotenv
import argparse  # Add this import
from datetime import datetime
import pandas as pd  # Add pandas for Excel export
import openpyxl  # Add openpyxl for Excel handling
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openai import AzureOpenAI

# Import utility functions for language detection and translation
from utils import (
    detect_document_language,
    translate_text,
    translate_document_pages,
    enhance_vision_prompt_for_arabic,
)

# Import token counting function from existing preprocessing
try:
    from a_preprocessing import count_tokens
except ImportError:
    # Fallback implementation if the import fails
    import tiktoken
    def count_tokens(text: str, model: str = "gpt-4o") -> int:
        """Count tokens in text using tiktoken"""
        try:
            encoding = tiktoken.encoding_for_model(model)
            tokens = encoding.encode(text)
            return len(tokens)
        except Exception as e:
            print(f"Token counting error: {e}")
            # Fallback estimation
            return len(text) // 4  # Rough estimate


# Load environment variables for API keys
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# Policy identification patterns
POLICY_PATTERNS = [
    # Law patterns
    r'Law No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})\s*(?:concerning|regarding|on|about)',
    r'Federal Law No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    r'Local Law No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    
    # Decree patterns
    r'Amiri Decree No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    r'Decree No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    r'Federal Decree No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    
    # Resolution patterns
    r'(?:Executive Council|Chairman of the Executive Council) Resolution No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    r'Cabinet Resolution No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    
    # Circular patterns
    r'Circular No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    
    # Arabic patterns (simplified, can be expanded)
    r'قانون رقم\s*\(\s*\d+\s*\)\s*لسنة\s*(\d{4})',
    r'مرسوم رقم\s*\(\s*\d+\s*\)\s*لسنة\s*(\d{4})',
    r'قرار رقم\s*\(\s*\d+\s*\)\s*لسنة\s*(\d{4})',
    r'تعميم رقم\s*\(\s*\d+\s*\)\s*لسنة\s*(\d{4})'
]

# Combined pattern for more efficient searching
COMBINED_POLICY_PATTERN = '|'.join(f'({pattern})' for pattern in POLICY_PATTERNS)

# TOC indicators - patterns that might indicate a table of contents
TOC_PATTERNS = [
    r"(?i)^[\s\*]*contents[\s\*]*$",
    r"(?i)^[\s\*]*table\s+of\s+contents[\s\*]*$",
    r"(?i)^[\s\*]*index[\s\*]*$",
    r"(?i)^[\s\*]*فهرس[\s\*]*$",
    r"(?i)^[\s\*]*المحتويات[\s\*]*$"
]

COMBINED_TOC_PATTERN = '|'.join(f'({pattern})' for pattern in TOC_PATTERNS)

# Page number pattern for finding page references in TOC
PAGE_NUMBER_PATTERN = r'\.{2,}|…+|\s{3,}|(\d+)\s*$'


# Pydantic models for structured data
class PageRange(BaseModel):
    """Model representing a page range"""
    start_page: int = Field(..., description="Starting page number (0-indexed)")
    end_page: Optional[int] = Field(None, description="Ending page number (0-indexed)")
    
    @property
    def page_list(self) -> List[int]:
        """Get list of pages in this range"""
        if self.end_page is None:
            return [self.start_page]
        return list(range(self.start_page, self.end_page + 1))


class PolicyEntry(BaseModel):
    """Model representing a policy entry from a table of contents or structure analysis"""
    name: str = Field(..., description="Name of the policy")
    page_range: PageRange = Field(..., description="Page range of the policy")
    confidence: float = Field(default=0.5, description="Confidence score of the extraction")


class DocStructure(BaseModel):
    """Model representing document structure analysis results"""
    toc_pages: List[int] = Field(default_factory=list, description="Pages containing table of contents")
    blank_pages: List[int] = Field(default_factory=list, description="Pages that are blank or near-blank")
    title_pages: List[int] = Field(default_factory=list, description="Pages that appear to be title pages")
    policies: List[PolicyEntry] = Field(default_factory=list, description="Policies found in the document")
    

class SmartChunk(BaseModel):
    """Model representing a smart chunk for processing"""
    chunk_id: int = Field(..., description="Unique ID for this chunk")
    policy_name: str = Field(..., description="Name of the policy this chunk belongs to")
    content: str = Field(..., description="Text content of the chunk")
    pages: List[int] = Field(..., description="Page numbers included in this chunk (0-indexed)")
    tokens: int = Field(..., description="Number of tokens in this chunk")
    complete_policy: bool = Field(default=False, description="Whether this chunk contains a complete policy")
    chunk_part: Optional[int] = Field(None, description="Part number for multi-part chunks")
    total_chunks: Optional[int] = Field(None, description="Total chunks for this policy")
    is_vision_extracted: bool = Field(default=False, description="Whether this chunk was extracted using vision")


class ChunkMetadata(BaseModel):
    """Model for chunk metadata"""
    total_chunks: int = Field(..., description="Total number of chunks")
    total_tokens: int = Field(..., description="Total tokens across all chunks")
    policies: List[str] = Field(..., description="List of policy names")
    chunk_summary: List[Dict[str, Any]] = Field(..., description="Summary of each chunk")


def is_blank_page(page_text: str) -> bool:
    """Check if a page is blank or near-blank"""
    # Remove whitespace
    cleaned_text = re.sub(r'\s+', '', page_text)
    # Consider blank if less than 50 non-whitespace characters
    return len(cleaned_text) < 50


def is_title_page(page_text: str) -> bool:
    """Check if a page is likely a title page"""
    # Title pages typically have limited text and specific patterns
    cleaned_text = page_text.strip()
    lines = [line.strip() for line in cleaned_text.split('\n') if line.strip()]
    
    # Title pages typically have few lines
    if len(lines) <= 5:
        # Look for year patterns
        year_pattern = r'\b(19|20)\d{2}\b'
        if re.search(year_pattern, cleaned_text):
            return True
            
    # Check for "cover" or equivalent words
    cover_words = ['cover', 'title', 'غلاف', 'عنوان']
    for word in cover_words:
        if word.lower() in cleaned_text.lower():
            return True
            
    return False


def page_to_image(page) -> Image.Image:
    """Convert a PDF page to a PIL Image"""
    zoom = 2  # Increase resolution
    matrix = fitz.Matrix(zoom, zoom)
    pixmap = page.get_pixmap(matrix=matrix)
    
    # Convert to PIL Image
    image = Image.frombytes("RGB", [pixmap.width, pixmap.height], pixmap.samples)
    return image


def setup_azure_openai_client():
    """Set up the Azure OpenAI client"""
    api_key = os.environ.get("Azure_OPENAI_API_KEY", "")
    endpoint = os.environ.get("Azure_OPENAI_ENDPOINT", "https://teams-ai-agent.openai.azure.com")
    api_version = "2024-10-21"
    
    client = AsyncAzureOpenAI(
        api_key=api_key,
        api_version=api_version,
        azure_endpoint=endpoint
    )
    
    return client


# Add a synchronous OpenAI client
def setup_sync_azure_openai_client():
    """Set up a synchronous Azure OpenAI client"""
    api_key = os.environ.get("Azure_OPENAI_API_KEY", "")
    endpoint = os.environ.get("Azure_OPENAI_ENDPOINT", "https://teams-ai-agent.openai.azure.com")
    api_version = "2024-10-21"
    
    client = AzureOpenAI(
        api_key=api_key,
        api_version=api_version,
        azure_endpoint=endpoint
    )
    
    return client


async def analyze_document_structure(doc_path: str, max_toc_pages: int = 5) -> DocStructure:
    """
    Analyze document structure to identify TOC, blank pages, title pages, and policies.
    This is a fallback method when vision analysis is not available.
    
    Args:
        doc_path: Path to the PDF document
        max_toc_pages: Maximum pages to consider for TOC analysis
        
    Returns:
        DocStructure object with analysis results
    """
    structure = DocStructure()
    
    try:
        doc = fitz.open(doc_path)
        toc_candidates = []
        
        # First pass: identify blank pages, title pages, and TOC candidates
        for page_num in range(min(max_toc_pages + 5, len(doc))):
            page = doc[page_num]
            page_text = page.get_text()
            
            # Check for blank pages
            if is_blank_page(page_text):
                structure.blank_pages.append(page_num)
                continue
                
            # Check for title pages
            if is_title_page(page_text):
                structure.title_pages.append(page_num)
                continue
                
            # Check for TOC indicators
            if any(re.search(pattern, page_text, re.MULTILINE) for pattern in TOC_PATTERNS):
                toc_candidates.append(page_num)
                structure.toc_pages.append(page_num)
                
                # Also check the next page as TOCs often span multiple pages
                if page_num + 1 < len(doc):
                    structure.toc_pages.append(page_num + 1)
        
        # Method 1: Use PDF's built-in TOC/outline if available
        pdf_toc = doc.get_toc()
        if pdf_toc:
            logger.info(f"Using built-in PDF TOC with {len(pdf_toc)} entries")
            for item in pdf_toc:
                level, title, page = item
                # Adjust page number (PDF TOC is 1-indexed)
                page_idx = page - 1
                
                # Create policy entry
                if re.search(COMBINED_POLICY_PATTERN, title, re.IGNORECASE):
                    policy = PolicyEntry(
                        name=title,
                        page_range=PageRange(start_page=page_idx, end_page=None),
                        confidence=0.85
                    )
                    structure.policies.append(policy)
        
        # Method 2: Parse TOC pages for policy entries
        if toc_candidates and (not pdf_toc or len(pdf_toc) < 3):
            logger.info(f"Parsing TOC pages: {toc_candidates}")
            
            for page_num in toc_candidates:
                page_text = doc[page_num].get_text()
                lines = page_text.split('\n')
                
                for i, line in enumerate(lines):
                    # Look for policy patterns in this line
                    if re.search(COMBINED_POLICY_PATTERN, line, re.IGNORECASE):
                        # Try to find page number at the end of the line
                        page_match = re.search(PAGE_NUMBER_PATTERN, line)
                        if page_match and page_match.group(1):
                            try:
                                target_page = int(page_match.group(1)) - 1  # Convert to 0-indexed
                                policy_name = line[:page_match.start()].strip()
                                
                                # Create policy entry
                                policy = PolicyEntry(
                                    name=policy_name,
                                    page_range=PageRange(start_page=target_page, end_page=None),
                                    confidence=0.75
                                )
                                structure.policies.append(policy)
                            except (ValueError, IndexError):
                                pass
        
        # Method 3: Text analysis of initial pages
        if not structure.policies or len(structure.policies) < 3:
            logger.info("Using text analysis of initial pages")
            
            # For simplicity, just use text analysis with patterns on first few pages
            for page_num in range(min(20, len(doc))):
                page_text = doc[page_num].get_text()
                lines = page_text.split('\n')
                
                for line in lines:
                    # Look for policy patterns
                    policy_match = re.search(COMBINED_POLICY_PATTERN, line, re.IGNORECASE)
                    if policy_match:
                        policy_name = line.strip()
                        
                        # Create policy entry
                        policy = PolicyEntry(
                            name=policy_name,
                            page_range=PageRange(start_page=page_num, end_page=None),
                            confidence=0.6
                        )
                        structure.policies.append(policy)
        
        # If no policies were found, create a single policy for the entire document
        if not structure.policies:
            logger.info("No policies found in document analysis, creating single policy for entire document")
            policy = PolicyEntry(
                name=os.path.basename(doc_path),
                page_range=PageRange(start_page=0, end_page=len(doc) - 1),
                confidence=0.3
            )
            structure.policies = [policy]
        
        # Second pass: estimate end pages for policies
        if len(structure.policies) > 1:
            for i in range(len(structure.policies) - 1):
                # End page of policy i is the start page of policy i+1 minus 1
                structure.policies[i].page_range.end_page = structure.policies[i+1].page_range.start_page - 1
            
            # Last policy ends at the last page of the document
            if structure.policies:
                last_policy = structure.policies[-1]
                # Skip blank pages at the end
                last_page = len(doc) - 1
                while last_page > last_policy.page_range.start_page and \
                      last_page in structure.blank_pages:
                    last_page -= 1
                last_policy.page_range.end_page = last_page
        
        # Validate page ranges to prevent extraction failures
        for policy in structure.policies:
            if policy.page_range.end_page is None:
                policy.page_range.end_page = policy.page_range.start_page
            
            # Ensure end page is not before start page
            if policy.page_range.end_page < policy.page_range.start_page:
                logger.warning(f"Invalid page range for {policy.name}: {policy.page_range.start_page+1}-{policy.page_range.end_page+1}")
                policy.page_range.end_page = policy.page_range.start_page
            
            # Ensure end page is within document bounds
            if policy.page_range.end_page >= len(doc):
                policy.page_range.end_page = len(doc) - 1
        
        # Close the document
        doc.close()
        
        return structure
        
    except Exception as e:
        logger.error(f"Error analyzing document structure: {e}")
        if 'doc' in locals():
            doc.close()
        return structure


def extract_policy_content(doc_path: str, policy: PolicyEntry, language_info: Dict[str, Any] = None, client = None) -> Dict[str, Any]:
    """
    Extract content for a specific policy
    
    Args:
        doc_path: Path to the PDF document
        policy: PolicyEntry object with policy information
        language_info: Optional language information
        client: Optional Azure OpenAI client for vision fallback
        
    Returns:
        Dict with extracted content and metadata
    """
    result = {
        "policy_name": policy.name,
        "content": "",
        "start_page": policy.page_range.start_page,
        "end_page": policy.page_range.end_page or policy.page_range.start_page,
        "is_vision_extracted": False
    }
    
    try:
        doc = fitz.open(doc_path)
        
        # Determine page range
        start_page = policy.page_range.start_page
        end_page = policy.page_range.end_page or start_page
        
        # Make sure we don't exceed document bounds
        end_page = min(end_page, len(doc) - 1)
        
        # Extract content from each page in the range
        content = []
        empty_pages = 0
        
        for page_num in range(start_page, end_page + 1):
            page = doc[page_num]
            page_text = page.get_text()
            
            if page_text.strip():
                content.append(f"[PAGE {page_num + 1}]\n{page_text}")
            else:
                empty_pages += 1
        
        # Join all content
        policy_content = "\n\n".join(content)
        
        # Add warning if no content was extracted
        if not content:
            logger.warning(f"No content extracted for policy: {policy.name} (pages {start_page+1}-{end_page+1})")
            
            # Check surrounding pages as a fallback (in case of off-by-one errors)
            logger.info(f"Checking surrounding pages for content...")
            surrounding_start = max(0, start_page - 1)
            surrounding_end = min(len(doc) - 1, end_page + 1)
            
            for page_num in range(surrounding_start, surrounding_end + 1):
                if page_num >= start_page and page_num <= end_page:
                    continue  # Skip already checked pages
                
                page = doc[page_num]
                page_text = page.get_text()
                
                if not is_blank_page(page_text):
                    logger.info(f"Found content on surrounding page {page_num+1}")
                    content.append(f"[PAGE {page_num + 1}]\n{page_text}")
            
            if content:
                policy_content = "\n\n".join(content)
                logger.info(f"Recovered content from surrounding pages for {policy.name}")
            elif client:  # Try vision extraction as a last resort
                logger.info(f"Attempting vision-based extraction for policy: {policy.name}")
                vision_content = extract_policy_content_with_vision(client, doc_path, start_page, end_page, language_info)
                if vision_content:
                    policy_content = vision_content
                    result["is_vision_extracted"] = True
                    logger.info(f"Successfully extracted content using vision for {policy.name}")
        
        # Handle language detection and translation
        is_arabic = False
        if language_info and language_info.get("is_arabic", False):
            is_arabic = True
            result["is_arabic"] = True
            
            # Only translate if we have content
            if policy_content.strip():
                try:
                    logger.info(f"Translating content for policy: {policy.name}")
                    # Store both original Arabic content and translated content
                    result["arabic_content"] = policy_content
                    result["content"] = translate_text(policy_content, source_lang='ar', target_lang='en')
                    logger.info(f"Successfully translated content ({len(policy_content)} chars)")
                except Exception as e:
                    logger.error(f"Error translating policy content: {e}")
                    # Use original content if translation fails
                    result["content"] = policy_content
            else:
                logger.warning(f"No content to translate for policy: {policy.name}")
        else:
            # For non-Arabic documents, just use the content as is
            result["content"] = policy_content
        
        # Close the document
        doc.close()
        
        return result
        
    except Exception as e:
        logger.error(f"Error extracting policy content: {e}")
        if 'doc' in locals():
            doc.close()
        return result


async def extract_policy_content_with_vision_async(client, doc_path, start_page, end_page, metadata=None):
    """
    Extract policy content using the vision API for pages where text extraction failed.
    This is an async version of the function that processes pages in small batches.
    
    Args:
        client: Azure OpenAI client (AsyncAzureOpenAI instance)
        doc_path: Path to the PDF document
        start_page: Start page of the policy (0-indexed)
        end_page: End page of the policy (0-indexed)
        metadata: Additional metadata about the policy
        
    Returns:
        Extracted content as string
    """
    is_arabic = metadata.get('is_arabic', False) if metadata else False
    
    try:
        doc = fitz.open(doc_path)
        batch_size = 2  # Process 2 pages at a time
        
        # Calculate number of batches
        total_pages = min(end_page - start_page + 1, len(doc) - start_page)
        num_batches = (total_pages + batch_size - 1) // batch_size  # Ceiling division
        
        all_extracted_content = []
        
        # Process each batch of pages
        for batch in range(num_batches):
            batch_start = start_page + (batch * batch_size)
            batch_end = min(batch_start + batch_size - 1, end_page, len(doc) - 1)
            
            logger.info(f"Processing vision extraction batch {batch+1}/{num_batches} (pages {batch_start+1}-{batch_end+1})")
            
            # Extract images from pages for this batch
            page_images = []
            for page_num in range(batch_start, batch_end + 1):
                if page_num >= len(doc):
                    break
                    
                page = doc[page_num]
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                img_bytes = pix.tobytes("png")
                img_base64 = base64.b64encode(img_bytes).decode('utf-8')
                
                page_images.append({
                    'page_num': page_num,
                    'base64': img_base64
                })
                
            # Skip if no images in this batch
            if not page_images:
                continue
                
            # Build vision prompt
            image_contents = []
            for page in page_images:
                image_contents.append({
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{page['base64']}",
                        "detail": "high"
                    }
                })
            
            prompt = f"""
            Extract ALL text content from these policy document pages. These are pages {batch_start+1} to {batch_end+1} of a policy document.
            
            Focus on:
            1. Policy title and number
            2. All article text in full detail
            3. Any tables, lists, or structured content
            4. Signatures, dates, and official notations
            
            For Arabic documents, capture the text in its original form without translation.
            
            Format the output as plain text, maintaining paragraph structure. Include markers like [PAGE X] to indicate page breaks.
            """
            
            # Make API call with all images - using await here
            response = await client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": f"You are an expert document text extractor specializing in extracting content from {'Arabic' if is_arabic else 'English'} policy documents, including text that might be difficult to read or embedded in images."},
                    {"role": "user", "content": [{"type": "text", "text": prompt}] + image_contents}
                ],
                max_tokens=4000
            )
            
            # Get the extracted content from this batch
            batch_content = response.choices[0].message.content
            logger.info(f"Vision extracted {len(batch_content)} chars from batch {batch+1}/{num_batches}")
            
            # Add to collected content
            all_extracted_content.append(batch_content)
        
        # Concatenate all extracted content
        extracted_content = "\n\n".join(all_extracted_content)
        logger.info(f"Total vision extraction: {len(extracted_content)} chars from {total_pages} pages")
        
        doc.close()
        return extracted_content
        
    except Exception as e:
        logger.error(f"Error in vision-based extraction: {e}")
        if 'doc' in locals():
            doc.close()
        return ""


def extract_policy_content_with_vision(client, doc_path, start_page, end_page, metadata=None):
    """
    Synchronous wrapper for extract_policy_content_with_vision_async.
    Use this when you need to call the vision extraction from synchronous code.
    Processes pages in small batches (2-3 pages) and concatenates the results.
    
    Args:
        client: Azure OpenAI client (the async client will be ignored and a new sync client created)
        doc_path: Path to the PDF document
        start_page: Start page of the policy (0-indexed)
        end_page: End page of the policy (0-indexed)
        metadata: Additional metadata about the policy
        
    Returns:
        Extracted content as string
    """
    # Create a synchronous client instead
    sync_client = setup_sync_azure_openai_client()
    
    # Extract is_arabic from metadata
    is_arabic = metadata.get('is_arabic', False) if metadata else False
    
    try:
        doc = fitz.open(doc_path)
        batch_size = 2  # Process 2 pages at a time
        
        # Calculate number of batches
        total_pages = min(end_page - start_page + 1, len(doc) - start_page)
        num_batches = (total_pages + batch_size - 1) // batch_size  # Ceiling division
        
        all_extracted_content = []
        
        # Process each batch of pages
        for batch in range(num_batches):
            batch_start = start_page + (batch * batch_size)
            batch_end = min(batch_start + batch_size - 1, end_page, len(doc) - 1)
            
            logger.info(f"Processing vision extraction batch {batch+1}/{num_batches} (pages {batch_start+1}-{batch_end+1})")
            
            # Extract images from pages for this batch
            page_images = []
            for page_num in range(batch_start, batch_end + 1):
                if page_num >= len(doc):
                    break
                    
                page = doc[page_num]
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                img_bytes = pix.tobytes("png")
                img_base64 = base64.b64encode(img_bytes).decode('utf-8')
                
                page_images.append({
                    'page_num': page_num,
                    'base64': img_base64
                })
                
            # Skip if no images in this batch
            if not page_images:
                continue
                
            # Build vision prompt
            image_contents = []
            for page in page_images:
                image_contents.append({
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{page['base64']}",
                        "detail": "high"
                    }
                })
            
            prompt = f"""
            Extract ALL text content from these policy document pages. These are pages {batch_start+1} to {batch_end+1} of a policy document.
            
            Focus on:
            1. Policy title and number
            2. All article text in full detail
            3. Any tables, lists, or structured content
            4. Signatures, dates, and official notations
            
            For Arabic documents, capture the text in its original form without translation.
            
            Format the output as plain text, maintaining paragraph structure. Include markers like [PAGE X] to indicate page breaks.
            """
            
            # Make API call with all images - using sync client
            response = sync_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": f"You are an expert document text extractor specializing in extracting content from {'Arabic' if is_arabic else 'English'} policy documents, including text that might be difficult to read or embedded in images."},
                    {"role": "user", "content": [{"type": "text", "text": prompt}] + image_contents}
                ],
                max_tokens=4000
            )
            
            # Get the extracted content from this batch
            batch_content = response.choices[0].message.content
            logger.info(f"Vision extracted {len(batch_content)} chars from batch {batch+1}/{num_batches}")
            
            # Add to collected content
            all_extracted_content.append(batch_content)
        
        # Concatenate all extracted content
        extracted_content = "\n\n".join(all_extracted_content)
        logger.info(f"Total vision extraction: {len(extracted_content)} chars from {total_pages} pages")
        
        doc.close()
        return extracted_content
        
    except Exception as e:
        logger.error(f"Error in vision-based extraction: {e}")
        if 'doc' in locals():
            doc.close()
        return ""


async def analyze_initial_pages_with_vision(client, doc_path, max_pages=5):
    """
    Analyze the first few pages of a document to identify TOC and policy boundaries
    
    Args:
        client: Azure OpenAI client
        doc_path: Path to the PDF document
        max_pages: Maximum number of initial pages to analyze
        
    Returns:
        Dict with TOC pages and identified policies
    """
    try:
        # First detect the document language
        lang_info = detect_document_language(doc_path, num_pages=min(max_pages + 5, 15))
        is_arabic = lang_info.get("is_arabic", False)
        doc_language = lang_info.get("dominant_language", "en")
        
        if is_arabic:
            logger.info(f"Detected Arabic document (confidence: {lang_info.get('confidence', 0):.2f})")
        else:
            logger.info(f"Detected document language: {doc_language}")
        
        doc = fitz.open(doc_path)
        page_images = []
        
        # Convert first few pages to images - use a higher number for TOC analysis
        for i in range(min(max_pages, len(doc))):
            page = doc[i]
            img = page_to_image(page)
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            base64_image = base64.b64encode(buffer.getvalue()).decode('utf-8')
            
            page_images.append({
                "page_num": i + 1,  # 1-indexed for human readability
                "base64": base64_image
            })
        
        doc.close()
        
        # Build message with multiple images
        image_contents = []
        for page in page_images:
            image_contents.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/png;base64,{page['base64']}",
                    "detail": "high"
                }
            })
        
        # Add text instruction before images
        prompt = f"""
        You are analyzing the first {len(page_images)} pages of a document to identify:
        
        1. Table of Contents (TOC) pages
        2. Policy titles and their starting page numbers
        3. The page number offset between PDF and printed page numbers
        4. Classification of TOC items as policies or non-policies
        
        These images are in sequence from page 1 to page {len(page_images)}. CAREFULLY EXAMINE ALL PROVIDED PAGES, as the table of contents may extend across multiple pages.
        
        CRITICAL INSTRUCTIONS: 
        - CAREFULLY review the entire table of contents to identify ALL policies and their page numbers.
        - EXAMINE EVERY PAGE provided for table of contents information - don't stop after the first TOC page.
        - Look for entries that represent laws, decrees, resolutions, or other legal documents.
        - Extract ALL policy names and their corresponding page numbers from these TOC pages.
        - Determine if there's an offset between PDF page numbers and printed page numbers.
        - For example, if the first page of content is labeled as page 1 in the document, 
          but it's actually the 2nd page in the PDF, the offset would be 1.
        
        DO NOT mark regular content pages as TOC pages.
        DO NOT extract policies from content pages - only from the TOC.
        EXTRACT EVERY POLICY LISTED IN THE TABLE OF CONTENTS - THIS IS CRITICAL.
        
        Format your response as JSON:
        {{
            "toc_pages": [page numbers containing table of contents],
            "policies": [
                {{
                    "policy_name": "Full policy name/title",
                    "page_number": starting page number as shown in the document,
                    "is_policy": true/false,
                    "policy_type": "law|decree|resolution|circular|other"
                }}
            ],
            "page_offset": the numerical offset between PDF page numbers and printed page numbers
        }}
        
        CLASSIFICATION RULES:
        - Mark as "is_policy: true" if the entry is a law, decree, resolution, or regulatory document
        - Examples of policies: "Law No. (5) of 2020", "Executive Council Resolution", "Decree No. (3)"
        - Mark as "is_policy: false" if the entry is a table of contents section, appendix, index, etc.
        
        If you can't determine the page offset with confidence, set it to 1 as a default.
        """
        
        # For Arabic documents, translate initial pages and enhance the prompt
        translated_text = ""
        if is_arabic:
            # Translate the first few pages to help with vision analysis
            translated_text = translate_document_pages(
                doc_path, 
                start_page=0, 
                end_page=min(max_pages, 4),
                source_lang='ar', 
                target_lang='en'
            )
            
            if translated_text:
                # First use the original enhance_vision_prompt_for_arabic function for consistent behavior
                prompt = enhance_vision_prompt_for_arabic(prompt, translated_text)
                
                # Add the additional instruction for capturing both original and translated names
                prompt += """
                
                EXTREMELY IMPORTANT:
                1. EXTRACT EVERY SINGLE POLICY from the table from images - don't miss any!
                2. For each policy, provide BOTH:
                   - The English translation in the "policy_name" field
                   - The original Arabic text in the "policy_name_arabic" field
                
                Your JSON output should include ALL policies with this format:
                {
                    "policy_name": "Law No. (4) of 2020",
                    "policy_name_arabic": "قانون رقم (4) لسنة 2020",
                    "page_number": 7,
                    "is_policy": true,
                    "policy_type": "law"
                }
                
                COMPLETENESS IS CRITICAL - CHECK MULTIPLE TIMES that you have extracted ALL policies from the table of contents.
                """
        
        # Add text content as first message part
        image_contents.insert(-1, {"type": "text", "text": prompt})
        
        # Make API call with all images
        response = await client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": f"You are an expert document analyzer specializing in identifying document structure, tables of contents, policy boundaries, and page numbering systems in legal and governmental documents. {'For Arabic documents, carefully review ALL pages provided to ensure you extract EVERY policy from the table of contents.' if is_arabic else ''}"},
                {"role": "user", "content": image_contents}
            ],
            response_format={"type": "json_object"}        )
        
        # Parse response
        result = response.choices[0].message.content
        logger.info(f"Received vision analysis result with {len(result)} characters")
        analysis = json.loads(result)
        
        # Ensure page_offset has a value, default to 1 if not provided
        if "page_offset" not in analysis or analysis["page_offset"] is None:
            analysis["page_offset"] = 1
            logger.info("Page offset not detected, using default offset of 1")
        else:
            logger.info(f"Detected page offset from vision: {analysis['page_offset']}")
            
        # Log the number of policies found
        if "policies" in analysis:
            logger.info(f"Vision analysis identified {len(analysis['policies'])} policies")
        else:
            logger.warning("No policies found in vision analysis result")
        
        # Process policies if we have an Arabic document
        if is_arabic and "policies" in analysis:
            for policy in analysis["policies"]:
                # If we don't have both policy_name and policy_name_arabic set correctly
                if "policy_name_arabic" not in policy and policy.get("policy_name"):
                    # The original name should be stored as policy_name_arabic
                    policy["policy_name_arabic"] = policy["policy_name"]
                    
                    # Translate if needed
                    if not policy.get("policy_name_translated"):
                        try:
                            policy["policy_name"] = translate_text(policy["policy_name"], "ar", "en")
                        except Exception as e:
                            logger.warning(f"Failed to translate policy name: {e}")
                            # Keep the original name if translation fails
                
                # Remove any redundant policy_name_translated field
                if "policy_name_translated" in policy:
                    # If we don't already have an English name, use the translated one
                    if policy.get("policy_name") == policy.get("policy_name_arabic"):
                        policy["policy_name"] = policy["policy_name_translated"]
                    
                    # Remove the redundant field
                    del policy["policy_name_translated"]
        
        # Add language information to analysis result
        analysis["language_info"] = {
            "is_arabic": is_arabic,
            "detected_language": doc_language,
            "confidence": lang_info.get("confidence", 0),
            "has_translation": bool(translated_text)
        }
        
        return analysis
        
    except Exception as e:
        logger.error(f"Error analyzing initial pages: {e}")
        return {"toc_pages": [], "policies": [], "page_offset": 1}


def create_doc_structure_from_vision(vision_analysis: Dict[str, Any], doc_path: str) -> DocStructure:
    """
    Convert vision analysis results to a DocStructure object
    
    Args:
        vision_analysis: Results from analyze_initial_pages_with_vision
        doc_path: Path to the PDF document
        
    Returns:
        DocStructure object with policies and page ranges
    """
    structure = DocStructure()
    
    try:
        doc = fitz.open(doc_path)
        total_pages = len(doc)
        
        # Get page offset from vision analysis (default to 1 if not detected)
        page_offset = vision_analysis.get("page_offset", 1)
        logger.info(f"Using page offset: {page_offset}")
        
        # Add TOC pages from vision analysis
        structure.toc_pages = [page - 1 for page in vision_analysis.get("toc_pages", [])]
        
        # Process policies from vision analysis
        if "policies" in vision_analysis and vision_analysis["policies"]:
            policies = vision_analysis["policies"]
            logger.info(f"Processing {len(policies)} policies from vision analysis for document structure")
            
            # Create policy entries from vision analysis
            for i, policy in enumerate(policies):
                try:
                    policy_name = policy.get("policy_name", "")
                    # Convert to 0-indexed page numbers and apply offset
                    doc_page_num = int(policy.get("page_number", 0))
                    pdf_page_num = (doc_page_num + page_offset) - 1  # Add offset and convert to 0-indexed
                    
                    if policy_name and pdf_page_num >= 0:
                        policy_entry = PolicyEntry(
                            name=policy_name,
                            page_range=PageRange(start_page=pdf_page_num, end_page=None),
                            confidence=0.95
                        )
                        structure.policies.append(policy_entry)
                        logger.debug(f"Added policy to structure: {policy_name} (page {pdf_page_num+1})")
                    else:
                        logger.warning(f"Invalid page number for policy: {policy_name}, page: {doc_page_num}, adjusted: {pdf_page_num}")
                except (ValueError, TypeError) as e:
                    logger.warning(f"Error processing vision policy entry: {e}")
            
            logger.info(f"Created {len(structure.policies)} policy entries from vision analysis")
            
            # Calculate end pages for policies
            if len(structure.policies) > 1:
                for i in range(len(structure.policies) - 1):
                    # End page of policy i is the start page of policy i+1 minus 1
                    structure.policies[i].page_range.end_page = structure.policies[i+1].page_range.start_page - 1
                
                # Last policy ends at the last page of the document
                if structure.policies:
                    last_policy = structure.policies[-1]
                    last_policy.page_range.end_page = total_pages - 1
        
        # If no policies found, create a single policy for the entire document
        if not structure.policies:
            logger.info("No policies found in vision analysis, creating single policy for entire document")
            policy = PolicyEntry(
                name=os.path.basename(doc_path),
                page_range=PageRange(start_page=0, end_page=total_pages - 1),
                confidence=0.3
            )
            structure.policies = [policy]
            
        # Identify blank and title pages for better extraction
        for page_num in range(min(20, total_pages)):
            page = doc[page_num]
            page_text = page.get_text()
            
            # Check for blank pages
            if is_blank_page(page_text):
                structure.blank_pages.append(page_num)
                continue
                
            # Check for title pages
            if is_title_page(page_text):
                structure.title_pages.append(page_num)
        
        # Validate page ranges to prevent extraction failures
        for policy in structure.policies:
            if policy.page_range.end_page is None:
                policy.page_range.end_page = policy.page_range.start_page
            
            # Ensure end page is not before start page
            if policy.page_range.end_page < policy.page_range.start_page:
                logger.warning(f"Invalid page range for {policy.name}: {policy.page_range.start_page+1}-{policy.page_range.end_page+1}")
                policy.page_range.end_page = policy.page_range.start_page
            
            # Ensure end page is within document bounds
            if policy.page_range.end_page >= total_pages:
                policy.page_range.end_page = total_pages - 1
        
        doc.close()
        logger.info(f"Final document structure has {len(structure.policies)} policies")
        return structure
    
    except Exception as e:
        logger.error(f"Error creating document structure from vision: {e}")
        if 'doc' in locals():
            doc.close()
        return structure


async def create_smart_chunks(doc_path: str, max_tokens: int = 100000, model: str = "gpt-4o", 
                              vision_analysis: Dict[str, Any] = None, client = None) -> List[Dict[str, Any]]:
    """
    Create smart chunks from a document based on policy boundaries
    
    Args:
        doc_path: Path to the PDF document
        max_tokens: Maximum tokens per chunk
        model: Model to use for token counting
        vision_analysis: Optional pre-computed vision analysis results
        client: Optional Azure OpenAI client for vision fallback
        
    Returns:
        List of dictionaries with chunk information
    """
    try:
        # Step 1: Get document structure (either from vision or analyze it now)
        if vision_analysis:
            logger.info(f"Using provided vision analysis for {doc_path}")
            doc_structure = create_doc_structure_from_vision(vision_analysis, doc_path)
        else:
            # Fallback to full analysis if no vision results provided
            logger.info(f"No vision analysis provided, analyzing document structure for {doc_path}")
            doc_structure = await analyze_document_structure(doc_path)
        
        if not doc_structure.policies:
            logger.warning("No policies found in document analysis")
            return []
        
        logger.info(f"Found {len(doc_structure.policies)} policies in document structure analysis")
        for policy in doc_structure.policies:
            logger.info(f"  Policy: {policy.name} (pages {policy.page_range.start_page+1}-{(policy.page_range.end_page or policy.page_range.start_page)+1})")
        
        # Get language information from vision analysis if available
        language_info = vision_analysis.get("language_info", {}) if vision_analysis else {}
        is_arabic = language_info.get("is_arabic", False)
        doc_language = language_info.get("detected_language", "en")
        
        # Step 2: Create chunks for each policy
        chunks = []
        
        for i, policy in enumerate(doc_structure.policies):
            logger.info(f"Processing policy {i+1}/{len(doc_structure.policies)}: {policy.name}")
            
            # Extract content for this policy
            content_result = extract_policy_content(doc_path, policy, language_info, client)
            content = content_result["content"]
            
            if not content:
                logger.warning(f"  Could not extract content for policy: {policy.name}")
                continue
                
            # Count tokens - use translated content for Arabic if available
            token_text = content if is_arabic and content else content
            token_count = count_tokens(token_text, model)
            logger.info(f"  Extracted {len(content)} chars, {token_count} tokens")
            
            # Check if we need to split into multiple chunks
            if token_count <= max_tokens:
                # Create a single chunk for this policy
                chunk = SmartChunk(
                    chunk_id=len(chunks) + 1,  # 1-indexed
                    policy_name=policy.name,
                    content=content,
                    pages=policy.page_range.page_list,
                    tokens=token_count,
                    complete_policy=True,
                    is_vision_extracted=content_result.get("is_vision_extracted", False)
                )
                chunks.append(chunk.dict())
            else:
                # Split content by pages and create multiple chunks
                logger.info(f"  Policy exceeds token limit ({token_count} > {max_tokens}), splitting into multiple chunks")
                
                # Get page markers
                page_splits = re.findall(r"\[PAGE (\d+)\]", content)
                if not page_splits:
                    # If no page markers, split by token count as a last resort
                    logger.warning(f"  No page markers found, splitting by token count")
                    # Simple method - split by paragraphs to avoid cutting mid-sentence
                    paragraphs = content.split("\n\n")
                    current_chunk = ""
                    current_tokens = 0
                    chunk_parts = []
                    
                    for para in paragraphs:
                        para_tokens = count_tokens(para, model)
                        if current_tokens + para_tokens > max_tokens:
                            if current_chunk:
                                chunk_parts.append(current_chunk)
                                current_chunk = para
                                current_tokens = para_tokens
                            else:
                                # Single paragraph exceeds limit, need to split it
                                logger.warning(f"  Single paragraph exceeds token limit ({para_tokens} tokens)")
                                chunk_parts.append(para)
                        else:
                            current_chunk += "\n\n" + para if current_chunk else para
                            current_tokens += para_tokens
                    
                    # Add the last chunk if not empty
                    if current_chunk:
                        chunk_parts.append(current_chunk)
                    
                    # Create a chunk for each part
                    for j, part in enumerate(chunk_parts):
                        part_tokens = count_tokens(part, model)
                        chunk = SmartChunk(
                            chunk_id=len(chunks) + 1,  # 1-indexed
                            policy_name=policy.name,
                            content=part,
                            pages=policy.page_range.page_list,
                            tokens=part_tokens,
                            complete_policy=False,
                            chunk_part=j + 1,
                            total_chunks=len(chunk_parts),
                            is_vision_extracted=content_result.get("is_vision_extracted", False)
                        )
                        chunks.append(chunk.dict())
                else:
                    # Split by pages for more granular control
                    page_contents = re.split(r"\[PAGE \d+\]", content)
                    page_contents = page_contents[1:] if len(page_contents) > 0 else page_contents  # Skip the first split which is empty
                    
                    # If that produces too few chunks, try paragraphs
                    if len(page_contents) <= 1:
                        logger.warning(f"  Page splitting insufficient, falling back to paragraph splitting")
                        # Fallback to paragraph splitting
                        paragraphs = content.split("\n\n")
                        current_chunk = ""
                        current_tokens = 0
                        chunk_parts = []
                        
                        for para in paragraphs:
                            para_tokens = count_tokens(para, model)
                            if current_tokens + para_tokens > max_tokens:
                                if current_chunk:
                                    chunk_parts.append(current_chunk)
                                    current_chunk = para
                                    current_tokens = para_tokens
                                else:
                                    # Single paragraph exceeds limit
                                    chunk_parts.append(para)
                            else:
                                current_chunk += "\n\n" + para if current_chunk else para
                                current_tokens += para_tokens
                        
                        # Add the last chunk if not empty
                        if current_chunk:
                            chunk_parts.append(current_chunk)
                        
                        # Create a chunk for each part
                        for j, part in enumerate(chunk_parts):
                            part_tokens = count_tokens(part, model)
                            chunk = SmartChunk(
                                chunk_id=len(chunks) + 1,  # 1-indexed
                                policy_name=policy.name,
                                content=part,
                                pages=policy.page_range.page_list,
                                tokens=part_tokens,
                                complete_policy=False,
                                chunk_part=j + 1,
                                total_chunks=len(chunk_parts),
                                is_vision_extracted=content_result.get("is_vision_extracted", False)
                            )
                            chunks.append(chunk.dict())
                    else:
                        # We can split by pages
                        current_chunk = ""
                        current_tokens = 0
                        current_pages = []
                        chunk_parts = []
                        chunk_page_lists = []
                        
                        for j, (page_text, page_num) in enumerate(zip(page_contents, page_splits)):
                            page_text = f"[PAGE {page_num}]\n{page_text}"
                            page_tokens = count_tokens(page_text, model)
                            page_num = int(page_num)
                            
                            if current_tokens + page_tokens > max_tokens:
                                if current_chunk:
                                    chunk_parts.append(current_chunk)
                                    chunk_page_lists.append(current_pages)
                                    current_chunk = page_text
                                    current_tokens = page_tokens
                                    current_pages = [page_num - 1]  # Store 0-indexed page
                                else:
                                    # Single page exceeds limit
                                    chunk_parts.append(page_text)
                                    chunk_page_lists.append([page_num - 1])
                            else:
                                current_chunk += "\n\n" + page_text if current_chunk else page_text
                                current_tokens += page_tokens
                                current_pages.append(page_num - 1)  # Store 0-indexed page
                        
                        # Add the last chunk if not empty
                        if current_chunk:
                            chunk_parts.append(current_chunk)
                            chunk_page_lists.append(current_pages)
                        
                        # Create a chunk for each part
                        for j, (part, pages) in enumerate(zip(chunk_parts, chunk_page_lists)):
                            part_tokens = count_tokens(part, model)
                            chunk = SmartChunk(
                                chunk_id=len(chunks) + 1,  # 1-indexed
                                policy_name=policy.name,
                                content=part,
                                pages=pages,
                                tokens=part_tokens,
                                complete_policy=False,
                                chunk_part=j + 1,
                                total_chunks=len(chunk_parts),
                                is_vision_extracted=content_result.get("is_vision_extracted", False)
                            )
                            chunks.append(chunk.dict())
        
        logger.info(f"Created {len(chunks)} smart chunks from {len(doc_structure.policies)} policies")
        return chunks
        
    except Exception as e:
        logger.error(f"Error creating smart chunks: {e}")
        return []


def save_policies_to_json(policies: List[Dict[str, Any]], output_dir: str, doc_path: str, analysis_result: Dict = None) -> None:
    """
    Save policies to disk in JSON format with enhanced metadata
    """
    # Get filename without extension
    base_filename = os.path.splitext(os.path.basename(doc_path))[0]
    
    # Create document-specific directory 
    doc_output_dir = os.path.join(output_dir, base_filename)
    
    # Clean up existing directory
    if os.path.exists(doc_output_dir):
        logger.info(f"Removing existing output directory: {doc_output_dir}")
        shutil.rmtree(doc_output_dir)
    
    os.makedirs(doc_output_dir, exist_ok=True)
    
    # Create policies directory
    policies_dir = os.path.join(doc_output_dir, "policies")
    os.makedirs(policies_dir, exist_ok=True)
    
    # Create document-level metadata
    doc_metadata = {
        "document": {
            "filename": os.path.basename(doc_path),
            "path": doc_path,
            "processing_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "language": analysis_result.get("language_info", {}).get("detected_language", "en"),
            "is_arabic": analysis_result.get("language_info", {}).get("is_arabic", False)
        },
        "policies": {
            "total_count": len(policies),
            "policy_count": sum(1 for p in policies if p.get("is_policy", True)),
            "non_policy_count": sum(1 for p in policies if not p.get("is_policy", True)),
            "vision_extracted_count": sum(1 for p in policies if p.get("metadata", {}).get("is_vision_extracted", False)),
            "summary": [
                {
                    "policy_id": policy.get("policy_id", i+1),  # Fallback to index if no ID
                    "policy_name": policy.get("policy_name", "Untitled Policy"),
                    "policy_name_arabic": policy.get("policy_name_arabic", ""),
                    "is_policy": policy.get("is_policy", True),
                    "page_range": policy.get("page_range", []),
                    "policy_type": policy.get("metadata", {}).get("policy_type", "unknown"),
                    "is_vision_extracted": policy.get("metadata", {}).get("is_vision_extracted", False)
                }
                for i, policy in enumerate(policies)
            ]
        }
    }
    
    # Save the vision analysis
    if analysis_result:
        doc_metadata["vision_analysis"] = analysis_result
        vision_file = os.path.join(doc_output_dir, "vision_analysis.json")
        with open(vision_file, "w") as f:
            json.dump(analysis_result, f, indent=2)
        logger.info(f"Saved vision analysis to {vision_file}")
    
    # Save document metadata
    with open(os.path.join(doc_output_dir, "document_metadata.json"), "w") as f:
        json.dump(doc_metadata, f, indent=2)
    
    # Save individual policy files
    for i, policy in enumerate(policies):
        policy_id = policy.get("policy_id", i+1)
        # Use English policy name for filenames
        policy_name = policy.get("policy_name", "Untitled Policy")
        policy_filename = f"policy_{policy_id:03d}_{clean_filename(policy_name[:50])}.json"
        with open(os.path.join(policies_dir, policy_filename), "w") as f:
            json.dump(policy, f, indent=2)
    
    logger.info(f"Saved {len(policies)} policies as JSON to {doc_output_dir}")


def policies_to_dataframe(policies: List[Dict[str, Any]], doc_path: str) -> pd.DataFrame:
    """
    Convert policy data to a pandas DataFrame for export to Excel
    
    Args:
        policies: List of policy dictionaries
        doc_path: Path to the source document
        
    Returns:
        DataFrame with policy data
    """
    # Extract data for each policy
    policy_data = []
    
    # Get relative path from current directory
    rel_path = os.path.relpath(doc_path)
    filename = os.path.basename(doc_path)
    
    for policy in policies:
        # Extract key metadata
        policy_id = policy.get("policy_id", 0)
        policy_name = policy.get("policy_name", "Untitled Policy")
        is_policy = policy.get("is_policy", True)
        doc_page = policy.get("document_page", 0)
        pdf_page = policy.get("pdf_page", 0)
        page_range = policy.get("page_range", [])
        page_range_str = ", ".join(map(str, page_range)) if page_range else ""
        policy_type = policy.get("metadata", {}).get("policy_type", "unknown")
        contains_tables = policy.get("metadata", {}).get("contains_tables", False)
        year = policy.get("metadata", {}).get("year", None)
        token_count = policy.get("metadata", {}).get("tokens", 0)
        
        # First ~200 characters of content as preview
        content_preview = policy.get("content", "")[:200] + "..." if policy.get("content") else ""
        
        # Add to data list
        policy_data.append({
            "document_path": rel_path,
            "filename": filename,
            "policy_id": policy_id,
            "policy_name": policy_name,
            "is_policy": is_policy,
            "policy_type": policy_type,
            "document_page": doc_page,
            "pdf_page": pdf_page,
            "page_range": page_range_str,
            "contains_tables": contains_tables,
            "year": year,
            "token_count": token_count,
            "content_preview": content_preview
        })
    
    # Create DataFrame
    df = pd.DataFrame(policy_data)
    
    # Reorder columns for better readability
    column_order = [
        "document_path", "filename", "policy_id", "policy_name", "is_policy", 
        "policy_type", "year", "document_page", "pdf_page", "page_range",
        "contains_tables", "token_count", "content_preview"
    ]
    
    # Ensure all columns exist (some might be missing in certain scenarios)
    existing_columns = [col for col in column_order if col in df.columns]
    df = df[existing_columns]
    
    return df


def save_policies_to_excel(policies: List[Dict[str, Any]], output_dir: str, doc_path: str) -> str:
    """
    Save policies to an Excel file for easy reviewing.
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Get document name without extension
    doc_name = os.path.splitext(os.path.basename(doc_path))[0]
    
    # Create subdirectory for this document
    doc_dir = os.path.join(output_dir, doc_name)
    os.makedirs(doc_dir, exist_ok=True)
    
    # Create Excel file path
    excel_path = os.path.join(doc_dir, f"{doc_name}_policies.xlsx")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default worksheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create summary worksheet
    summary_sheet = wb.create_sheet("Summary")
    
    # Add headers
    summary_headers = [
        "Policy ID", "Policy Name", "Arabic Name", "Is Policy", "Document Page", "PDF Page",
        "Page Range", "Language", "Is Arabic", "Year", "Policy Type", "Contains Tables", 
        "Contains Images", "Tokens", "Relative Path"
    ]
    summary_sheet.append(summary_headers)
    
    # Set header style
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
    
    # Add policy data
    for policy in policies:
        metadata = policy.get("metadata", {})
        
        # Get English policy name for filename (already translated if Arabic)
        policy_name = policy.get("policy_name", "")
        
        # Calculate relative path to individual policy file - use English name for filename
        rel_path = f"./policies/policy_{policy.get('policy_id', 0):03d}_{clean_filename(policy_name[:50])}.json"
        
        # Format page range as a string
        page_range = ', '.join(map(str, policy.get('page_range', [])))
        
        summary_row = [
            policy.get("policy_id"),
            policy_name,
            policy.get("policy_name_arabic", ""),  # Include Arabic name if available
            policy.get("is_policy", True),
            policy.get("document_page"),
            policy.get("pdf_page"),
            page_range,
            metadata.get("language", ""),
            metadata.get("is_arabic", False),
            metadata.get("year"),
            metadata.get("policy_type", "unknown"),
            metadata.get("contains_tables", False),
            metadata.get("contains_images", False),
            metadata.get("tokens", 0),
            rel_path
        ]
        summary_sheet.append(summary_row)
    
    # Adjust column widths
    for col in summary_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 80 else 80
        summary_sheet.column_dimensions[column].width = adjusted_width
    
    # Create English content sheet
    content_sheet = wb.create_sheet("Content")
    
    # Add headers for content sheet
    content_headers = [
        "Policy ID", "Policy Name", "Is Policy", "Content",
    ]
    content_sheet.append(content_headers)
    
    # Set header style
    for cell in content_sheet[1]:
        cell.font = Font(bold=True)
    
    # Add policy content (always the main content - translated for Arabic documents)
    for policy in policies:
        content_row = [
            policy.get("policy_id"),
            policy.get("policy_name"),
            policy.get("is_policy", True),
            policy.get("content", "")
        ]
        content_sheet.append(content_row)
    
    # Adjust column widths
    for col in content_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if column != 'D':  # Skip the content column
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
        if column != 'D':  # Skip the content column
            adjusted_width = (max_length + 2) if max_length < 80 else 80
            content_sheet.column_dimensions[column].width = adjusted_width
    
    # Set content column width
    content_sheet.column_dimensions['D'].width = 120
    
    # Create Arabic content sheet if needed
    has_arabic = any(policy.get("policy_name_arabic") or policy.get("arabic_content") for policy in policies)
    if has_arabic:
        arabic_sheet = wb.create_sheet("Arabic Content")
        
        # Add headers for Arabic sheet
        arabic_headers = [
            "Policy ID", "English Name", "Arabic Name", "Arabic Content",
        ]
        arabic_sheet.append(arabic_headers)
        
        # Set header style
        for cell in arabic_sheet[1]:
            cell.font = Font(bold=True)
        
        # Add Arabic policy content
        for policy in policies:
            if policy.get("policy_name_arabic") or policy.get("arabic_content"):
                arabic_row = [
                    policy.get("policy_id"),
                    policy.get("policy_name"),
                    policy.get("policy_name_arabic", ""),
                    policy.get("arabic_content", "")
                ]
                arabic_sheet.append(arabic_row)
        
        # Adjust column widths
        for col in arabic_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if column != 'D':  # Skip the content column
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            if column != 'D':  # Skip the content column
                adjusted_width = (max_length + 2) if max_length < 80 else 80
                arabic_sheet.column_dimensions[column].width = adjusted_width
        
        # Set content column width
        arabic_sheet.column_dimensions['D'].width = 120
    
    # Save workbook
    wb.save(excel_path)
    logger.info(f"Saved policy information to Excel file: {excel_path}")
    
    return excel_path


def append_to_master_excel(source_excel_path, master_excel_path):
    """Append policy data from a document Excel file to the master Excel file."""
    import pandas as pd
    from openpyxl import load_workbook
    
    # Try to read the source Excel file
    try:
        source_wb = load_workbook(source_excel_path)
        source_sheet = source_wb.get_sheet_by_name("Summary")
    except:
        logger.error(f"Failed to read source Excel file: {source_excel_path}")
        return
    
    # Convert source sheet to DataFrame
    source_data = []
    
    # Get headers from first row
    headers = [cell.value for cell in source_sheet[1]]
    
    # Get data from remaining rows
    for row in source_sheet.iter_rows(min_row=2):
        row_data = [cell.value for cell in row]
        source_data.append(dict(zip(headers, row_data)))
    
    source_df = pd.DataFrame(source_data)
    
    # Add source document information
    doc_name = os.path.basename(os.path.dirname(source_excel_path))
    source_df["Document"] = doc_name
    
    # Update relative path to include document folder
    if "Relative Path" in source_df.columns:
        source_df["Relative Path"] = source_df["Relative Path"].apply(
            lambda x: f"./{doc_name}/{x.lstrip('./')}" if isinstance(x, str) else x
        )
    
    # Check if master Excel file exists
    if os.path.exists(master_excel_path):
        # Load existing master file
        try:
            master_df = pd.read_excel(master_excel_path)
            
            # Append source data to master data
            master_df = pd.concat([master_df, source_df], ignore_index=True)
        except:
            logger.warning(f"Failed to read master Excel file. Creating new one.")
            master_df = source_df
    else:
        # Create new master Excel file
        master_df = source_df
    
    # Save master Excel file
    try:
        master_df.to_excel(master_excel_path, index=False)
        logger.info(f"Updated master Excel file: {master_excel_path}")
    except Exception as e:
        logger.error(f"Failed to save master Excel file: {e}")


def clean_filename(filename: str) -> str:
    """Create a clean, safe filename from a string"""
    # Replace non-alphanumeric characters with underscores
    clean = re.sub(r'[^\w\s-]', '_', filename)
    # Replace multiple spaces with single underscore
    clean = re.sub(r'\s+', '_', clean)
    # Limit length
    if len(clean) > 50:
        clean = clean[:47] + "..."
    return clean


def detect_tables_in_content(content: str) -> bool:
    """Detect if content likely contains tables based on patterns"""
    # Simple detection based on patterns
    table_patterns = [
        r"\|\s*\w+\s*\|",  # |  Text  |
        r"\+[-+]+\+",      # +-----+-----+
        r"[^\|]+\|[^\|]+\|[^\|]+",  # Text | Text | Text
        r"[-]{3,}\s*\n.*\n[-]{3,}"  # ---- Header ----
    ]
    
    for pattern in table_patterns:
        if re.search(pattern, content):
            return True
    
    return False


async def main():
    """Main function for CLI usage"""
    parser = argparse.ArgumentParser(description='Smart document chunking for policy extraction')
    parser.add_argument('--input', '-i', required=True, help='Path to PDF document')
    parser.add_argument('--output', '-o', default='output', help='Output directory for chunks')
    parser.add_argument('--max-tokens', '-m', type=int, default=100000, help='Maximum tokens per chunk')
    parser.add_argument('--model', type=str, default='gpt-4o', help='Model for token counting')
    parser.add_argument('--toc-pages', type=int, default=5, help='Number of pages to analyze for table of contents')
    parser.add_argument('--no-excel', action='store_true', help='Skip Excel file generation')
    parser.add_argument('--no-translation', action='store_true', help='Skip translation for Arabic documents')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    
    args = parser.parse_args()
    
    # Set debug logging if requested
    if args.debug:
        logger.setLevel(logging.DEBUG)
        logger.debug("Debug logging enabled")
    
    # Check if input file exists
    if not os.path.exists(args.input):
        logger.error(f"Input file not found: {args.input}")
        return
    
    logger.info(f"Processing document: {args.input}")
    
    # Set up Azure OpenAI client for analysis
    client = setup_azure_openai_client()
    
    # Analyze the initial pages to detect policies - this is now the ONLY vision analysis
    logger.info(f"Analyzing document structure with vision using {args.toc_pages} pages...")
    analysis_result = await analyze_initial_pages_with_vision(
        client, args.input, max_pages=args.toc_pages
    )
    
    # Get language information
    language_info = analysis_result.get("language_info", {})
    is_arabic = language_info.get("is_arabic", False)
    doc_language = language_info.get("detected_language", "en")
    
    if is_arabic:
        if args.no_translation:
            logger.info("Arabic document detected, but translation is disabled")
        else:
            logger.info(f"Arabic document detected (confidence: {language_info.get('confidence', 0):.2f})")
    else:
        logger.info(f"Document language: {doc_language}")
    
    # Log policy information from vision analysis
    if "policies" in analysis_result and analysis_result["policies"]:
        page_offset = analysis_result.get("page_offset", 1)
        policies = analysis_result["policies"]
        logger.info(f"Found {len(policies)} policies in table of contents with page offset: {page_offset}")
        
        for i, policy in enumerate(policies):
            # Correctly show PDF page calculation: document page + offset (1-indexed for display)
            doc_page = int(policy.get("page_number", 0))
            pdf_page = doc_page + page_offset
            policy_name = policy.get("policy_name", "")
            policy_name_arabic = policy.get("policy_name_arabic", "")
            
            if is_arabic and policy_name_arabic:
                logger.info(f"  Policy {i+1}: {policy_name} (Arabic: {policy_name_arabic}) (document page {doc_page}, PDF page {pdf_page})")
            else:
                logger.info(f"  Policy {i+1}: {policy_name} (document page {doc_page}, PDF page {pdf_page})")
    else:
        logger.warning("No policies found in vision analysis")
    
    # Create smart chunks - passing the vision analysis results directly
    logger.info("Creating smart chunks from document structure...")
    chunks = await create_smart_chunks(args.input, args.max_tokens, args.model, analysis_result, client)
    logger.info(f"Created {len(chunks)} smart chunks")
    
    # Transform vision analysis policies to full policy objects
    if "policies" in analysis_result and analysis_result["policies"]:
        # Create a mapping of policy name to chunks
        policy_chunks = {}
        for chunk in chunks:
            policy_name = chunk["policy_name"]
            if policy_name not in policy_chunks:
                policy_chunks[policy_name] = []
            policy_chunks[policy_name].append(chunk)
        
        logger.info(f"Mapping chunks to {len(policy_chunks)} policies")
        
        # Create full policy objects
        full_policies = []
        missing_policies = []
        
        for i, policy_data in enumerate(analysis_result["policies"]):
            # Get policy name (already translated if Arabic)
            policy_name = policy_data.get("policy_name", "")
            policy_name_arabic = policy_data.get("policy_name_arabic", "")
            
            # Find matching chunks
            policy_content_chunks = policy_chunks.get(policy_name, [])
            
            if not policy_content_chunks:
                logger.warning(f"No chunks found for policy: {policy_name}")
                missing_policies.append(policy_name)
                continue
            
            # Extract year from policy name if possible
            year_match = re.search(r'(?:of|لسنة)\s*(\d{4})', policy_name if not policy_name_arabic else policy_name_arabic)
            year = int(year_match.group(1)) if year_match else None
            
            # Calculate page range
            doc_page = int(policy_data.get("page_number", 0))
            pdf_page = doc_page + page_offset - 1  # 0-indexed
            
            # Get the pages from chunks or use a default
            pages = []
            for chunk in policy_content_chunks:
                pages.extend(chunk.get("pages", []))
            
            if not pages and doc_page > 0:
                # Default to just the starting page if no chunks
                pages = [pdf_page]
            
            # Check for tables in content and gather content
            has_tables = False
            full_content = ""
            
            for chunk in policy_content_chunks:
                chunk_content = chunk.get("content", "")
                full_content += chunk_content + "\n\n"
                
                if detect_tables_in_content(chunk_content):
                    has_tables = True
            
            # Create policy object
            full_policy = {
                "policy_id": i + 1,  # 1-indexed ID for user friendliness
                "policy_name": policy_name,
                "is_policy": policy_data.get("is_policy", True),
                "document_page": doc_page,
                "pdf_page": doc_page + page_offset,  # 1-indexed for display
                "page_range": sorted(list(set(pages))),
                "content": full_content.strip() if full_content else "",
                "content_chunks": [
                    {
                        "chunk_id": chunk.get("chunk_id", j+1),
                        "content": chunk.get("content", ""),
                        "pages": chunk.get("pages", [])
                    }
                    for j, chunk in enumerate(policy_content_chunks)
                ],
                "metadata": {
                    "contains_tables": has_tables,
                    "contains_images": False,  # Not detecting images yet
                    "language": doc_language,
                    "is_arabic": is_arabic,
                    "year": year,
                    "policy_type": policy_data.get("policy_type", "unknown"),
                    "tokens": sum(chunk.get("tokens", 0) for chunk in policy_content_chunks),
                    "is_vision_extracted": any(chunk.get("is_vision_extracted", False) for chunk in policy_content_chunks)
                }
            }
            
            # If content is empty or minimal, try vision extraction
            if not full_content.strip() or len(full_content.strip()) < 200:
                logger.info(f"No or minimal content found for policy {i+1}: {policy_name}. Attempting vision extraction...")
                start_page = min(pages) if pages else pdf_page
                end_page = max(pages) if pages else pdf_page
                
                # Convert from 1-indexed to 0-indexed for the extraction function
                start_page = start_page - 1 if start_page > 0 else 0
                end_page = end_page - 1 if end_page > 0 else 0
                
                vision_content = extract_policy_content_with_vision(
                    client, 
                    args.input, 
                    start_page, 
                    end_page, 
                    {"is_arabic": is_arabic}
                )
                
                if vision_content and len(vision_content.strip()) > len(full_content.strip()):
                    logger.info(f"Successfully extracted content using vision for policy {i+1}: {policy_name}")
                    full_policy["content"] = vision_content
                    full_policy["metadata"]["is_vision_extracted"] = True
                    
                    # For Arabic documents, store the original content
                    if is_arabic:
                        full_policy["arabic_content"] = vision_content
                        
                        # Try to translate the content
                        if not args.no_translation:
                            try:
                                logger.info(f"Translating vision-extracted content for policy: {policy_name}")
                                full_policy["content"] = translate_text(vision_content, source_lang='ar', target_lang='en')
                                logger.info(f"Successfully translated vision-extracted content ({len(vision_content)} chars)")
                            except Exception as e:
                                logger.error(f"Error translating vision-extracted content: {e}")
                                # Keep the Arabic content if translation fails
                                full_policy["content"] = vision_content
                else:
                    logger.warning(f"Vision extraction failed or didn't improve content for policy {i+1}: {policy_name}")
            
            # Add Arabic policy name if available (for Arabic documents)
            if is_arabic and policy_name_arabic:
                full_policy["policy_name_arabic"] = policy_name_arabic
            
            # Add original Arabic content if available and this is an Arabic document
            if is_arabic and len(policy_content_chunks) > 0:
                first_chunk = policy_content_chunks[0]
                if "arabic_content" in first_chunk:
                    # Only keep the Arabic content at the policy level, not in chunks
                    full_policy["arabic_content"] = first_chunk["arabic_content"]
            
            full_policies.append(full_policy)
        
        if missing_policies:
            logger.warning(f"Could not find content for {len(missing_policies)} policies: {', '.join(missing_policies[:5])}" + 
                          (f" and {len(missing_policies) - 5} more" if len(missing_policies) > 5 else ""))
        
        # Save policies and analysis to disk as JSON
        if full_policies:
            logger.info(f"Saving {len(full_policies)} policies to disk")
            save_policies_to_json(full_policies, args.output, args.input, analysis_result)
            logger.info(f"Successfully saved {len(full_policies)} policies to {args.output}")
            
            # Generate Excel files if not disabled
            if not args.no_excel:
                # Save individual Excel file for this document
                excel_path = save_policies_to_excel(full_policies, args.output, args.input)
                
                # Update master Excel file
                append_to_master_excel(excel_path, os.path.join(args.output, "all_policies.xlsx"))
                logger.info(f"Added policies to master Excel file: {os.path.join(args.output, 'all_policies.xlsx')}")
        else:
            logger.warning("No policies created. Document may not contain policies or chunking failed.")
    else:
        logger.warning("No policies found in vision analysis")


if __name__ == "__main__":
    asyncio.run(main()) 