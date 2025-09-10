#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Smart Chunking Module for Policy Extraction

This module provides advanced document chunking strategies specifically designed
for policy extraction from large documents. It includes:

- Policy structure analysis from table of contents and initial pages
- Image-based analysis of initial pages for policy identification
- Smart document segmentation based on policy identification
- Automated blank and title page removal
- Efficient policy boundary detection
"""

import re
import os
import io
import base64
import logging
import shutil  # Add import for directory removal
from typing import List, Dict, Any, Optional, Tuple, Set, Union
import fitz  # PyMuPDF
import json
import tempfile
from dataclasses import dataclass, field
from collections import defaultdict
from pydantic import BaseModel, Field
from PIL import Image
from openai import AsyncAzureOpenAI
import asyncio
from io import BytesIO
from dotenv import load_dotenv
import argparse  # Add this import
from datetime import datetime
import pandas as pd  # Add pandas for Excel export
import openpyxl  # Add openpyxl for Excel handling
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openai import AzureOpenAI
from concurrent.futures import ThreadPoolExecutor, as_completed

# Import utility functions from utils.py
from utils import (
    detect_document_language,
    translate_text,
    translate_document_pages,
    enhance_vision_prompt_for_arabic,
    is_blank_page,
    is_title_page,
    page_to_image,
    clean_filename,
    detect_tables_in_content,
    setup_azure_openai_client,
    has_tables_in_pdf,     # For the memory-efficient check
    detect_tables_in_pdf,  # Add this function import
    is_valid_table,
)

# Import token counting function from existing preprocessing
try:
    from a_preprocessing import count_tokens
except ImportError:
    # Fallback implementation if the import fails
    import tiktoken
    def count_tokens(text: str, model: str = "gpt-5") -> int:
        """Count tokens in text using tiktoken"""
        try:
            encoding = tiktoken.encoding_for_model(model)
            tokens = encoding.encode(text)
            return len(tokens)
        except Exception as e:
            print(f"Token counting error: {e}")
            # Fallback estimation
            return len(text) // 4  # Rough estimate


# Load environment variables for API keys
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
# Concurrency controls via environment
VISION_BATCH_CONCURRENCY = int(os.environ.get("VISION_BATCH_CONCURRENCY", "2"))
TABLE_VISION_CONCURRENCY = int(os.environ.get("TABLE_VISION_CONCURRENCY", "3"))

def _encode_page_to_jpeg_b64(page, scale: float = 1.8, quality: int = 70) -> str:
    """
    Convert a PDF page to a base64-encoded JPEG for efficient vision requests.
    """
    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    buf = BytesIO()
    img.save(buf, format="JPEG", quality=quality, optimize=True)
    return base64.b64encode(buf.getvalue()).decode('utf-8')


# Policy identification patterns
POLICY_PATTERNS = [
    # Law patterns
    r'Law No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})\s*(?:concerning|regarding|on|about)',
    r'Federal Law No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    r'Local Law No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    
    # Decree patterns
    r'Amiri Decree No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    r'Decree No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    r'Federal Decree No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    
    # Resolution patterns
    r'(?:Executive Council|Chairman of the Executive Council) Resolution No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    r'Cabinet Resolution No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    
    # Circular patterns
    r'Circular No\.\s*\(\s*\d+\s*\)\s*of\s*(?:Year\s*)?(\d{4})',
    
    # Arabic patterns (simplified, can be expanded)
    r'قانون رقم\s*\(\s*\d+\s*\)\s*لسنة\s*(\d{4})',
    r'مرسوم رقم\s*\(\s*\d+\s*\)\s*لسنة\s*(\d{4})',
    r'قرار رقم\s*\(\s*\d+\s*\)\s*لسنة\s*(\d{4})',
    r'تعميم رقم\s*\(\s*\d+\s*\)\s*لسنة\s*(\d{4})'
]

# Combined pattern for more efficient searching
COMBINED_POLICY_PATTERN = '|'.join(f'({pattern})' for pattern in POLICY_PATTERNS)

# TOC indicators - patterns that might indicate a table of contents
TOC_PATTERNS = [
    r"(?i)^[\s\*]*contents[\s\*]*$",
    r"(?i)^[\s\*]*table\s+of\s+contents[\s\*]*$",
    r"(?i)^[\s\*]*index[\s\*]*$",
    r"(?i)^[\s\*]*فهرس[\s\*]*$",
    r"(?i)^[\s\*]*المحتويات[\s\*]*$"
]

COMBINED_TOC_PATTERN = '|'.join(f'({pattern})' for pattern in TOC_PATTERNS)

# Page number pattern for finding page references in TOC
PAGE_NUMBER_PATTERN = r'\.{2,}|…+|\s{3,}|(\d+)\s*$'


# Pydantic models for structured data
class PageRange(BaseModel):
    """Model representing a page range"""
    start_page: int = Field(..., description="Starting page number (0-indexed)")
    end_page: Optional[int] = Field(None, description="Ending page number (0-indexed)")
    
    @property
    def page_list(self) -> List[int]:
        """Get list of pages in this range"""
        if self.end_page is None:
            return [self.start_page]
        return list(range(self.start_page, self.end_page + 1))


class PolicyEntry(BaseModel):
    """Model representing a policy entry from a table of contents or structure analysis"""
    name: str = Field(..., description="Name of the policy")
    page_range: PageRange = Field(..., description="Page range of the policy")
    confidence: float = Field(default=0.5, description="Confidence score of the extraction")


class DocStructure(BaseModel):
    """Model representing document structure analysis results"""
    toc_pages: List[int] = Field(default_factory=list, description="Pages containing table of contents")
    blank_pages: List[int] = Field(default_factory=list, description="Pages that are blank or near-blank")
    title_pages: List[int] = Field(default_factory=list, description="Pages that appear to be title pages")
    policies: List[PolicyEntry] = Field(default_factory=list, description="Policies found in the document")
    

class SmartChunk(BaseModel):
    """Model representing a smart chunk for processing"""
    chunk_id: int = Field(..., description="Unique ID for this chunk")
    policy_name: str = Field(..., description="Name of the policy this chunk belongs to")
    content: str = Field(..., description="Text content of the chunk")
    pages: List[int] = Field(..., description="Page numbers included in this chunk (0-indexed)")
    tokens: int = Field(..., description="Number of tokens in this chunk")
    complete_policy: bool = Field(default=False, description="Whether this chunk contains a complete policy")
    chunk_part: Optional[int] = Field(None, description="Part number for multi-part chunks")
    total_chunks: Optional[int] = Field(None, description="Total chunks for this policy")
    is_vision_extracted: bool = Field(default=False, description="Whether this chunk was extracted using vision")


class ChunkMetadata(BaseModel):
    """Model for chunk metadata"""
    total_chunks: int = Field(..., description="Total number of chunks")
    total_tokens: int = Field(..., description="Total tokens across all chunks")
    policies: List[str] = Field(..., description="List of policy names")
    chunk_summary: List[Dict[str, Any]] = Field(..., description="Summary of each chunk")


def extract_policy_content(doc_path: str, policy: PolicyEntry, language_info: Dict[str, Any] = None, client = None, skip_vision_for_tables: bool = False) -> Dict[str, Any]:
    """
    Extract content for a specific policy, including detecting and extracting tables
    
    Args:
        doc_path: Path to the PDF document
        policy: PolicyEntry object with policy information
        language_info: Optional language information
        client: Optional Azure OpenAI client for vision fallback
        skip_vision_for_tables: If True, skip using vision model for tables extraction
        
    Returns:
        Dict with extracted content and metadata
    """
    result = {
        "policy_name": policy.name,
        "content": "",
        "start_page": policy.page_range.start_page,
        "end_page": policy.page_range.end_page or policy.page_range.start_page,
        "is_vision_extracted": False,
        "tables": [],  # Add tables field to store extracted tables
        "contains_tables": False  # Flag to indicate if tables were found
    }
    
    try:
        doc = fitz.open(doc_path)
        
        # Determine page range
        start_page = policy.page_range.start_page
        end_page = policy.page_range.end_page or start_page
        
        # Make sure we don't exceed document bounds
        end_page = min(end_page, len(doc) - 1)
        
        # Extract content from each page in the range
        content = []
        empty_pages = 0
        
        # Determine if we need RTL handling for this text
        is_arabic = False
        if language_info and language_info.get("is_arabic", False):
            is_arabic = True
        
        # STEP 1: First, check if the policy pages contain any tables (memory-efficient)
        page_numbers = list(range(start_page, end_page + 1))
        try:
            contains_tables = has_tables_in_pdf(doc_path, page_numbers)
        except Exception:
            contains_tables = False
        
        if contains_tables:
            logger.info(f"Tables detected in policy: {policy.name} (pages {start_page+1}-{end_page+1})")
            result["contains_tables"] = True
            
            # STEP 2: If we need to extract content with vision model, do it for the whole pages
            # with tables directly without the intermediate detection step
            if client and not skip_vision_for_tables:
                # Simplified approach: Let vision model handle table detection and extraction
                # by sending the entire pages with tables
                metadata = {"is_arabic": is_arabic, "contains_tables": True}
                
                vision_content = extract_policy_content_with_vision(
                    client,
                    doc_path,
                    start_page,
                    end_page,
                    metadata
                )
                
                if vision_content:
                    logger.info(f"Successfully extracted content with tables using vision model")
                    result["content"] = vision_content
                    result["is_vision_extracted"] = True
                    
                    # For Arabic, add original content from vision extraction
                    if is_arabic and metadata.get("full_arabic_content"):
                        result["arabic_content"] = metadata["full_arabic_content"]
                    
                    # Close the document since we're using vision extraction
                    doc.close()
                    return result
            elif skip_vision_for_tables:
                logger.info(f"Skipping vision model for tables in policy: {policy.name} (tables detected)")
        
        # STEP 3: Extract regular text content from each page (if not already extracted with vision)
        for page_num in range(start_page, end_page + 1):
            page = doc[page_num]
            page_text = page.get_text()
            
            if page_text.strip():
                # Add page marker and text
                content.append(f"[PAGE {page_num + 1}]\n{page_text}")
            else:
                empty_pages += 1
        
        # Join all content
        policy_content = "\n\n".join(content)
        
        # STEP 4: Handle empty or minimal content cases
        if not content:
            logger.warning(f"No content extracted for policy: {policy.name} (pages {start_page+1}-{end_page+1})")
            
            # Check surrounding pages as a fallback (in case of off-by-one errors)
            logger.info("Checking surrounding pages for content...")
            surrounding_start = max(0, start_page - 1)
            surrounding_end = min(len(doc) - 1, end_page + 1)
            
            for page_num in range(surrounding_start, surrounding_end + 1):
                if page_num >= start_page and page_num <= end_page:
                    continue  # Skip already checked pages
                
                page = doc[page_num]
                page_text = page.get_text()
                
                if not is_blank_page(page_text):
                    logger.info(f"Found content on surrounding page {page_num+1}")
                    content.append(f"[PAGE {page_num + 1}]\n{page_text}")
            
            if content:
                policy_content = "\n\n".join(content)
                logger.info(f"Recovered content from surrounding pages for {policy.name}")
            elif client:  # Try vision extraction as a last resort
                logger.info(f"Attempting vision-based extraction for policy: {policy.name}")
                vision_metadata = {"is_arabic": is_arabic}
                vision_content = extract_policy_content_with_vision(client, doc_path, start_page, end_page, vision_metadata)
                
                if vision_content:
                    policy_content = vision_content
                    result["is_vision_extracted"] = True
                    logger.info(f"Successfully extracted content using vision for {policy.name}")
                    
                    # Check if the vision extraction found tables
                    if vision_metadata.get("contains_tables", False):
                        result["contains_tables"] = True
        
        # STEP 5: Handle language detection and translation
        if is_arabic:
            result["is_arabic"] = True
            
            # Only translate if we have content
            if policy_content.strip():
                try:
                    logger.info(f"Translating content for policy: {policy.name}")
                    # Store both original Arabic content and translated content
                    result["arabic_content"] = policy_content
                    result["content"] = translate_text(policy_content, source_lang='ar', target_lang='en')
                    logger.info(f"Successfully translated content ({len(policy_content)} chars)")
                except Exception as e:
                    logger.error(f"Error translating policy content: {e}")
                    # Use original content if translation fails
                    result["content"] = policy_content
            else:
                logger.warning(f"No content to translate for policy: {policy.name}")
        else:
            # For non-Arabic documents, just use the content as is
            result["content"] = policy_content
        
        # Close the document
        doc.close()
        
        return result
        
    except Exception as e:
        logger.error(f"Error extracting policy content: {e}")
        if 'doc' in locals():
            doc.close()
        return result


def extract_policy_content_with_vision(client, doc_path, start_page, end_page, metadata=None):
    """
    Extract policy content using the vision API for pages where text extraction failed.
    Processes pages in small batches and concatenates the results.
    For Arabic documents, handles both extraction and translation in a single step.
    
    Args:
        client: Azure OpenAI client (unused, a new sync client will be created)
        doc_path: Path to the PDF document
        start_page: Start page of the policy (0-indexed)
        end_page: End page of the policy (0-indexed)
        metadata: Additional metadata about the policy
        
    Returns:
        Extracted content as string
    """
    # Create a synchronous client
    sync_client = setup_azure_openai_client(async_client=False)
    
    # Extract is_arabic from metadata
    is_arabic = metadata.get('is_arabic', False) if metadata else False
    contains_tables = metadata.get('contains_tables', False) if metadata else False
    
    try:
        doc = fitz.open(doc_path)
        batch_size = 3  # conservative default per batch
        total_pages = min(end_page - start_page + 1, len(doc) - start_page)
        num_batches = (total_pages + batch_size - 1) // batch_size

        def build_and_call_batch(batch_index: int):
            b_start = start_page + (batch_index * batch_size)
            b_end = min(b_start + batch_size - 1, end_page, len(doc) - 1)
            logger.info(f"Vision extracting batch {batch_index+1}/{num_batches} (pages {b_start+1}-{b_end+1})")
            page_images = []
            for page_num in range(b_start, b_end + 1):
                if page_num >= len(doc):
                    break
                page = doc[page_num]
                img_b64 = _encode_page_to_jpeg_b64(page, scale=1.8, quality=70)
                page_images.append({'page_num': page_num, 'base64': img_b64})
            if not page_images:
                return "", False, None
            image_contents = []
            for p in page_images:
                image_contents.append({
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/jpeg;base64,{p['base64']}",
                        "detail": "high"
                    }
                })
            # Create a prompt that handles both extraction and translation if needed
            # Include specific instructions for tables if tables are detected
            table_instructions = ""
            if contains_tables:
                table_instructions = """
                IMPORTANT TABLE HANDLING INSTRUCTIONS:
                - These pages contain one or more tables
                - Extract ALL tables completely and accurately
                - Format tables as properly structured markdown tables with aligned columns
                - Include column headers and maintain the original structure
                - Tables should be formatted like this:
                  | Header 1 | Header 2 | Header 3 |
                  | -------- | -------- | -------- |
                  | Data 1   | Data 2   | Data 3   |
                  | Data 4   | Data 5   | Data 6   |
                - Ensure all cells and data are properly extracted
                - Maintain the context and position of tables within the document
                """
            
            if is_arabic:
                prompt = f"""
                Extract ALL text content from these policy document pages. These are pages {batch_start+1} to {batch_end+1} of a policy document.
                
                The document is in Arabic. Please:
                1. First extract the Arabic text in its original form
                2. Then translate it to English with high accuracy
                
                For the Arabic extraction:
                - Maintain the RIGHT-TO-LEFT reading order
                - Extract text line by line in the EXACT order it appears visually
                - Preserve the original RTL formatting
                
                For the English translation:
                - Provide an accurate translation that preserves legal terminology
                - Format the translation in standard English paragraph structure
                
                {table_instructions}
                
                Format your response in two sections:
                
                [ARABIC CONTENT]
                (Original Arabic text here, including tables in markdown format)
                
                [ENGLISH TRANSLATION]
                (English translation here, including tables in markdown format)
                
                Include markers like [PAGE X] in both sections to indicate page breaks.
                """
            else:
                prompt = f"""
                Extract ALL text content from these policy document pages. These are pages {batch_start+1} to {batch_end+1} of a policy document.
                
                Focus on:
                1. Policy title and number
                2. All article text in full detail
                3. Any tables, lists, or structured content
                4. Signatures, dates, and official notations
                
                {table_instructions}
                
                Format the output as plain text, maintaining paragraph structure. Include markers like [PAGE X] to indicate page breaks.
                """
            
            # Make API call with all images - using sync client
            response = sync_client.chat.completions.create(
                model="gpt-5",
                messages=[
                    {"role": "system", "content": f"You are an expert document text extractor specializing in extracting content from {'Arabic' if is_arabic else 'English'} policy documents, including text that might be difficult to read or embedded in images. Pay special attention to properly extracting tables in a structured markdown format."},
                    {"role": "user", "content": [{"type": "text", "text": prompt}] + image_contents}
                ],
                max_completion_tokens=15000,
                seed=42
            )
            
            # Get the extracted content from this batch
            batch_content = response.choices[0].message.content
            
            # Check if the batch contains tables by looking for markdown table syntax
            table_indicators = [
                "|--", "| --", "|-", "| -",   # Header separator formats
                "| --- |", "|---|",           # Centered header separators
                "| Header |", "| Column |"    # Explicit header texts
            ]
            
            has_tables = any(indicator in batch_content for indicator in table_indicators)
            arabic_piece = None
            if is_arabic and "[ENGLISH TRANSLATION]" in batch_content:
                parts = batch_content.split("[ENGLISH TRANSLATION]")
                if len(parts) > 1:
                    arabic_piece = parts[0].replace("[ARABIC CONTENT]", "").strip()
                    batch_content = parts[1].strip()
            return batch_content, has_tables, arabic_piece

        all_extracted_content = []
        any_tables = False
        collected_arabic = []

        # Parallelize batches with a bounded thread pool
        with ThreadPoolExecutor(max_workers=VISION_BATCH_CONCURRENCY) as ex:
            futures = {ex.submit(build_and_call_batch, b): b for b in range(num_batches)}
            for fut in as_completed(futures):
                content, has_tables, arabic_piece = fut.result()
                if content:
                    all_extracted_content.append(content)
                if has_tables:
                    any_tables = True
                if arabic_piece:
                    collected_arabic.append(arabic_piece)

        if any_tables and metadata is not None:
            metadata["contains_tables"] = True
        
        logger.info(f"Extracted {sum(len(c) for c in all_extracted_content)} total chars across {len(all_extracted_content)} batches")
        
        # Concatenate all extracted content
        extracted_content = "\n\n".join(all_extracted_content)
        
        # For Arabic, concatenate all Arabic content pieces
        if is_arabic and metadata is not None and collected_arabic:
            metadata["full_arabic_content"] = "\n\n".join(collected_arabic)
        
        doc.close()
        return extracted_content
        
    except Exception as e:
        logger.error(f"Error in vision-based extraction: {e}")
        if 'doc' in locals():
            doc.close()
        return ""


async def analyze_initial_pages_with_vision(client, doc_path, max_pages=2):
    """
    Analyze the first few pages of a document to identify TOC and policy boundaries
    
    Args:
        client: Azure OpenAI client
        doc_path: Path to the PDF document
        max_pages: Maximum number of initial pages to analyze
        
    Returns:
        Dict with TOC pages and identified policies
    """
    try:
        # First detect the document language
        lang_info = detect_document_language(doc_path, num_pages=min(max_pages + 5, 15))
        is_arabic = lang_info.get("is_arabic", False)
        doc_language = lang_info.get("dominant_language", "en")
        
        if is_arabic:
            logger.info(f"Detected Arabic document (confidence: {lang_info.get('confidence', 0):.2f})")
        else:
            logger.info(f"Detected document language: {doc_language}")
        
        doc = fitz.open(doc_path)
        total_pages = len(doc)
        page_images = []
        
        # Convert first few pages to images - use a higher number for TOC analysis
        for i in range(min(max_pages, total_pages)):
            page = doc[i]
            img = page_to_image(page)
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            base64_image = base64.b64encode(buffer.getvalue()).decode('utf-8')
            
            page_images.append({
                "page_num": i + 1,  # 1-indexed for human readability
                "base64": base64_image
            })
        
        doc.close()
        
        # Build message with multiple images
        image_contents = []
        for page in page_images:
            image_contents.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/png;base64,{page['base64']}",
                    "detail": "high"
                }
            })
        
        # Add text instruction before images
        prompt = f"""
        You are analyzing the first {len(page_images)} pages of a document to identify:
        
        1. Table of Contents (TOC) pages
        2. Policy titles and their starting page numbers
        3. The page number offset between PDF and printed page numbers
        4. Classification of TOC items as policies or non-policies
        5. Document title and main policy title if no TOC is present
        
        These images are in sequence from page 1 to page {len(page_images)}. CAREFULLY EXAMINE ALL PROVIDED PAGES.
        
        IMPORTANT: This document has a total of {total_pages} pages. Ensure all page numbers you reference are within this range.
        
        CRITICAL INSTRUCTIONS:
        - First, determine if the document has a table of contents (TOC)
        - If a TOC exists:
          - CAREFULLY review the entire table of contents to identify ALL policies and their page numbers
          - EXAMINE EVERY PAGE provided for table of contents information - don't stop after the first TOC page
          - Look for entries that represent laws, decrees, resolutions, or other legal documents
          - Extract ALL policy names and their corresponding page numbers from these TOC pages
        - If NO table of contents is found:
          - Look for document titles and policy titles in the initial pages
          - If you find MULTIPLE distinct policy titles, list each one as a separate policy
          - If you find only ONE title, assume the entire document is about this single policy
          - If no clear title is found, use the document filename as the policy name
        - Determine if there's an offset between PDF page numbers and printed page numbers
        - For example, if the first page of content is labeled as page 1 in the document, 
          but it's actually the 2nd page in the PDF, the offset would be 1
        
        DO NOT mark regular content pages as TOC pages.
        EXTRACT EVERY POLICY LISTED IN THE TABLE OF CONTENTS - THIS IS CRITICAL.
        
        Format your response as JSON:
        {{
            "toc_pages": [page numbers containing table of contents],
            "has_toc": true/false,
            "policies": [
                {{
                    "policy_name": "Full policy name/title",
                    "page_number": starting page number as shown in the document,
                    "is_policy": true/false,
                    "policy_type": "law|decree|resolution|circular|other"
                }}
            ],
            "page_offset": the numerical offset between PDF page numbers and printed page numbers
        }}
        
        CLASSIFICATION RULES:
        - Mark as "is_policy: true" if the entry is a law, decree, resolution, or regulatory document
        - Examples of policies: "Law No. (5) of 2020", "Executive Council Resolution", "Decree No. (3)"
        - Mark as "is_policy: false" if the entry is a table of contents section, appendix, index, etc.
        
        If you can't determine the page offset with confidence, set it to 1 as a default.
        """
        
        # For Arabic documents, translate initial pages and enhance the prompt
        translated_text = ""
        if is_arabic:
            # Translate the first few pages to help with vision analysis
            translated_text = translate_document_pages(
                doc_path, 
                start_page=0, 
                end_page=min(max_pages, 4),
                source_lang='ar', 
                target_lang='en'
            )
            
            if translated_text:
                # First use the original enhance_vision_prompt_for_arabic function for consistent behavior
                prompt = enhance_vision_prompt_for_arabic(prompt, translated_text)
                
                # Add the additional instruction for capturing both original and translated names
                prompt += """
                
                EXTREMELY IMPORTANT:
                1. EXTRACT EVERY SINGLE POLICY from the table of contents - don't miss any!
                2. For each policy, provide BOTH:
                   - The English translation in the "policy_name" field
                   - The original Arabic text in the "policy_name_arabic" field
                
                Your JSON output should include ALL policies with this format:
                {
                    "policy_name": "Law No. (4) of 2020",
                    "policy_name_arabic": "قانون رقم (4) لسنة 2020",
                    "page_number": 7,
                    "is_policy": true,
                    "policy_type": "law"
                }
                
                COMPLETENESS IS CRITICAL - CHECK MULTIPLE TIMES that you have extracted ALL policies from the table of contents.
                
                IF THE DOCUMENT HAS NO TABLE OF CONTENTS:
                1. Look for document titles in the initial pages
                2. If you find multiple distinct policy titles, list each one as a separate policy
                3. If you find only one title, assume the entire document is about this single policy
                4. For each policy identified from titles, provide both Arabic and English names
                """
        
        # Add text content as first message part
        image_contents.insert(0, {"type": "text", "text": prompt})
        
        # Make API call with all images
        response = await client.chat.completions.create(
            model="gpt-5",
            messages=[
                {"role": "system", "content": f"You are an expert document analyzer specializing in identifying document structure, tables of contents, policy boundaries, and page numbering systems in legal and governmental documents. {'For Arabic documents, carefully review ALL pages provided to ensure you extract EVERY policy from the table of contents.' if is_arabic else ''} If a document has no table of contents, carefully identify the document title and any policy titles from the initial pages - if only one title is found, assume the entire document is about this single policy."},
                {"role": "user", "content": image_contents}
            ],
            response_format={"type": "json_object"},
            max_completion_tokens=15000,
            seed=42
        )
        
        # Parse response
        result = response.choices[0].message.content
        logger.info(f"Received vision analysis result with {len(result)} characters")
        analysis = json.loads(result)
        
        # Ensure page_offset has a value, default to 1 if not provided
        if "page_offset" not in analysis or analysis["page_offset"] is None:
            analysis["page_offset"] = 1
            logger.info("Page offset not detected, using default offset of 1")
        else:
            logger.info(f"Detected page offset from vision: {analysis['page_offset']}")
        
        # Log whether a table of contents was found
        has_toc = analysis.get("has_toc", False) 
        if has_toc:
            toc_pages = analysis.get("toc_pages", [])
            logger.info(f"Document has a table of contents on pages: {toc_pages}")
        else:
            logger.info("Document does not have a table of contents, processing based on document titles")
            
        # Log the number of policies found
        if "policies" in analysis:
            logger.info(f"Vision analysis identified {len(analysis['policies'])} policies")
        else:
            logger.warning("No policies found in vision analysis result")
        
        # Process policies if we have an Arabic document
        if is_arabic and "policies" in analysis:
            for policy in analysis["policies"]:
                # If we don't have both policy_name and policy_name_arabic set correctly
                if "policy_name_arabic" not in policy and policy.get("policy_name"):
                    # The original name should be stored as policy_name_arabic
                    policy["policy_name_arabic"] = policy["policy_name"]
                    
                    # Translate if needed
                    if not policy.get("policy_name_translated"):
                        try:
                            policy["policy_name"] = translate_text(policy["policy_name"], "ar", "en")
                        except Exception as e:
                            logger.warning(f"Failed to translate policy name: {e}")
                            # Keep the original name if translation fails
                
                # Remove any redundant policy_name_translated field
                if "policy_name_translated" in policy:
                    # If we don't already have an English name, use the translated one
                    if policy.get("policy_name") == policy.get("policy_name_arabic"):
                        policy["policy_name"] = policy["policy_name_translated"]
                    
                    # Remove the redundant field
                    del policy["policy_name_translated"]
        
        # Add language information to analysis result
        analysis["language_info"] = {
            "is_arabic": is_arabic,
            "detected_language": doc_language,
            "confidence": lang_info.get("confidence", 0),
            "has_translation": bool(translated_text),
            "has_toc": analysis.get("has_toc", False)
        }
        
        return analysis
        
    except Exception as e:
        logger.error(f"Error analyzing initial pages: {e}")
        return {"toc_pages": [], "policies": [], "page_offset": 1}


def create_doc_structure_from_vision(vision_analysis: Dict[str, Any], doc_path: str) -> DocStructure:
    """
    Convert vision analysis results to a DocStructure object
    
    Args:
        vision_analysis: Results from analyze_initial_pages_with_vision
        doc_path: Path to the PDF document
        
    Returns:
        DocStructure object with policies and page ranges
    """
    structure = DocStructure()
    
    try:
        doc = fitz.open(doc_path)
        total_pages = len(doc)
        logger.info(f"Document has {total_pages} total pages")
        
        # Get page offset from vision analysis (default to 1 if not detected)
        page_offset = vision_analysis.get("page_offset", 1)
        logger.info(f"Using page offset: {page_offset}")
        
        # Add TOC pages from vision analysis
        structure.toc_pages = [page - 1 for page in vision_analysis.get("toc_pages", [])]
        
        # Process policies from vision analysis
        if "policies" in vision_analysis and vision_analysis["policies"]:
            policies = vision_analysis["policies"]
            logger.info(f"Processing {len(policies)} policies from vision analysis for document structure")
            
            # Create policy entries from vision analysis
            for i, policy in enumerate(policies):
                try:
                    policy_name = policy.get("policy_name", "")
                    
                    # Get page number with proper error handling
                    try:
                        doc_page_num = int(policy.get("page_number", 0))
                    except (ValueError, TypeError):
                        logger.warning(f"Invalid page number format for policy: {policy_name}, using page 1")
                        doc_page_num = 1
                    
                    # Convert to 0-indexed page numbers and apply offset
                    pdf_page_num = (doc_page_num + page_offset) - 1  # Add offset and convert to 0-indexed
                    
                    # Ensure page number is within document bounds
                    if pdf_page_num >= total_pages:
                        logger.warning(f"Policy page number {pdf_page_num+1} exceeds document page count {total_pages}, using first page")
                        pdf_page_num = 0
                    
                    if policy_name and pdf_page_num >= 0:
                        policy_entry = PolicyEntry(
                            name=policy_name,
                            page_range=PageRange(start_page=pdf_page_num, end_page=None),
                            confidence=0.95
                        )
                        structure.policies.append(policy_entry)
                        logger.debug(f"Added policy to structure: {policy_name} (page {pdf_page_num+1})")
                    else:
                        logger.warning(f"Invalid page number for policy: {policy_name}, page: {doc_page_num}, adjusted: {pdf_page_num}")
                except Exception as e:
                    logger.warning(f"Error processing vision policy entry: {e}")
            
            logger.info(f"Created {len(structure.policies)} policy entries from vision analysis")
            
            # Calculate end pages for policies
            if len(structure.policies) > 1:
                for i in range(len(structure.policies) - 1):
                    # End page of policy i is the start page of policy i+1 minus 1
                    structure.policies[i].page_range.end_page = structure.policies[i+1].page_range.start_page - 1
                
                # Last policy ends at the last page of the document
                if structure.policies:
                    last_policy = structure.policies[-1]
                    last_policy.page_range.end_page = total_pages - 1
        
        # If no policies found, create a single policy for the entire document
        if not structure.policies:
            logger.info("No policies found in vision analysis, creating single policy for entire document")
            policy = PolicyEntry(
                name=os.path.basename(doc_path),
                page_range=PageRange(start_page=0, end_page=total_pages - 1),
                confidence=0.3
            )
            structure.policies = [policy]
            
        # Identify blank and title pages for better extraction
        for page_num in range(min(20, total_pages)):
            page = doc[page_num]
            page_text = page.get_text()
            
            # Check for blank pages
            if is_blank_page(page_text):
                structure.blank_pages.append(page_num)
                continue
                
            # Check for title pages
            if is_title_page(page_text):
                structure.title_pages.append(page_num)
        
        # Validate page ranges to prevent extraction failures
        for policy in structure.policies:
            # Ensure start page is within bounds
            if policy.page_range.start_page >= total_pages:
                logger.warning(f"Policy start page {policy.page_range.start_page+1} exceeds document page count {total_pages}")
                policy.page_range.start_page = 0
            
            # Set end page to start page if not set
            if policy.page_range.end_page is None:
                policy.page_range.end_page = min(policy.page_range.start_page + 5, total_pages - 1)
                logger.info(f"Setting end page for {policy.name} to {policy.page_range.end_page+1}")
            
            # Ensure end page is not before start page
            if policy.page_range.end_page < policy.page_range.start_page:
                logger.warning(f"Invalid page range for {policy.name}: {policy.page_range.start_page+1}-{policy.page_range.end_page+1}, fixing")
                policy.page_range.end_page = min(policy.page_range.start_page + 5, total_pages - 1)
            
            # Ensure end page is within document bounds
            if policy.page_range.end_page >= total_pages:
                logger.warning(f"Policy end page {policy.page_range.end_page+1} exceeds document page count {total_pages}, adjusting")
                policy.page_range.end_page = total_pages - 1
            
            # Log final page range for this policy
            logger.info(f"Final page range for policy '{policy.name}': pages {policy.page_range.start_page+1}-{policy.page_range.end_page+1}")
        
        doc.close()
        logger.info(f"Final document structure has {len(structure.policies)} policies")
        return structure
    
    except Exception as e:
        logger.error(f"Error creating document structure from vision: {e}")
        if 'doc' in locals():
            doc.close()
        return structure


async def create_smart_chunks(doc_path: str, max_tokens: int = 50000, model: str = "gpt-5", 
                              vision_analysis: Dict[str, Any] = None, client = None, 
                              skip_vision_for_tables: bool = False) -> List[Dict[str, Any]]:
    """
    Create smart chunks from a document based on policy boundaries
    
    Args:
        doc_path: Path to the PDF document
        max_tokens: Maximum tokens per chunk
        model: Model to use for token counting
        vision_analysis: Optional pre-computed vision analysis results
        client: Optional Azure OpenAI client for vision fallback
        skip_vision_for_tables: If True, skip using vision model for tables extraction
        
    Returns:
        List of dictionaries with chunk information
    """
    try:
        # Step 1: Get document structure (either from vision or analyze it now)
        if vision_analysis:
            logger.info(f"Using provided vision analysis for {doc_path}")
            doc_structure = create_doc_structure_from_vision(vision_analysis, doc_path)
        
        if not doc_structure.policies:
            logger.warning("No policies found in document analysis")
            return []
        
        logger.info(f"Found {len(doc_structure.policies)} policies in document structure analysis")
        for policy in doc_structure.policies:
            logger.info(f"  Policy: {policy.name} (pages {policy.page_range.start_page+1}-{(policy.page_range.end_page or policy.page_range.start_page)+1})")
        
        # Get language information from vision analysis if available
        language_info = vision_analysis.get("language_info", {}) if vision_analysis else {}
        is_arabic = language_info.get("is_arabic", False)
        doc_language = language_info.get("detected_language", "en")
        
        # Step 2: Create chunks for each policy
        chunks = []
        
        for i, policy in enumerate(doc_structure.policies):
            logger.info(f"Processing policy {i+1}/{len(doc_structure.policies)}: {policy.name}")
            
            # Extract content for this policy
            content_result = extract_policy_content(
                doc_path, 
                policy, 
                language_info, 
                client,
                skip_vision_for_tables
            )
            content = content_result["content"]
            
            if not content:
                logger.warning(f"  Could not extract content for policy: {policy.name}")
                continue
                
            # Count tokens - use translated content for Arabic if available
            token_text = content if is_arabic and content else content
            token_count = count_tokens(token_text, model)
            logger.info(f"  Extracted {len(content)} chars, {token_count} tokens")
            
            # Check if we need to split into multiple chunks
            if token_count <= max_tokens:
                # Create a single chunk for this policy
                chunk = SmartChunk(
                    chunk_id=len(chunks) + 1,  # 1-indexed
                    policy_name=policy.name,
                    content=content,
                    pages=policy.page_range.page_list,
                    tokens=token_count,
                    complete_policy=True,
                    is_vision_extracted=content_result.get("is_vision_extracted", False)
                )
                chunk_dict = chunk.dict()
                
                # Add tables and table flag if present
                if content_result.get("contains_tables", False):
                    chunk_dict["contains_tables"] = True
                if "tables" in content_result and content_result["tables"]:
                    chunk_dict["tables"] = content_result["tables"]
                
                # Add Arabic content if available
                if is_arabic and "arabic_content" in content_result:
                    chunk_dict["arabic_content"] = content_result["arabic_content"]
                
                chunks.append(chunk_dict)
            else:
                # Split content by pages and create multiple chunks
                logger.info(f"  Policy exceeds token limit ({token_count} > {max_tokens}), splitting into multiple chunks")
                
                # Get page markers
                page_splits = re.findall(r"\[PAGE (\d+)\]", content)
                if not page_splits:
                    # If no page markers, split by token count as a last resort
                    logger.warning(f"  No page markers found, splitting by token count")
                    # Simple method - split by paragraphs to avoid cutting mid-sentence
                    paragraphs = content.split("\n\n")
                    current_chunk = ""
                    current_tokens = 0
                    chunk_parts = []
                    
                    for para in paragraphs:
                        para_tokens = count_tokens(para, model)
                        if current_tokens + para_tokens > max_tokens:
                            if current_chunk:
                                chunk_parts.append(current_chunk)
                                current_chunk = para
                                current_tokens = para_tokens
                            else:
                                # Single paragraph exceeds limit, need to split it
                                logger.warning(f"  Single paragraph exceeds token limit ({para_tokens} tokens)")
                                chunk_parts.append(para)
                        else:
                            current_chunk += "\n\n" + para if current_chunk else para
                            current_tokens += para_tokens
                    
                    # Add the last chunk if not empty
                    if current_chunk:
                        chunk_parts.append(current_chunk)
                    
                    # Create a chunk for each part
                    for j, part in enumerate(chunk_parts):
                        part_tokens = count_tokens(part, model)
                        chunk = SmartChunk(
                            chunk_id=len(chunks) + 1,  # 1-indexed
                            policy_name=policy.name,
                            content=part,
                            pages=policy.page_range.page_list,
                            tokens=part_tokens,
                            complete_policy=False,
                            chunk_part=j + 1,
                            total_chunks=len(chunk_parts),
                            is_vision_extracted=content_result.get("is_vision_extracted", False)
                        )
                        chunk_dict = chunk.dict()
                        
                        # Add tables and table flag if present - but only to the first chunk to avoid duplication
                        if j == 0:
                            if content_result.get("contains_tables", False):
                                chunk_dict["contains_tables"] = True
                            if "tables" in content_result and content_result["tables"]:
                                chunk_dict["tables"] = content_result["tables"]
                        
                        # Add Arabic content if available
                        if is_arabic and "arabic_content" in content_result:
                            # Need to split Arabic content similarly - this is a simplification
                            # In a real implementation, we should split Arabic content by paragraphs too
                            chunk_dict["arabic_content"] = content_result["arabic_content"]
                        
                        chunks.append(chunk_dict)
                else:
                    # Split by pages for more granular control
                    page_contents = re.split(r"\[PAGE \d+\]", content)
                    page_contents = page_contents[1:] if len(page_contents) > 0 else page_contents  # Skip the first split which is empty
                    
                    # If that produces too few chunks, try paragraphs
                    if len(page_contents) <= 1:
                        logger.warning(f"  Page splitting insufficient, falling back to paragraph splitting")
                        # Fallback to paragraph splitting
                        paragraphs = content.split("\n\n")
                        current_chunk = ""
                        current_tokens = 0
                        chunk_parts = []
                        
                        for para in paragraphs:
                            para_tokens = count_tokens(para, model)
                            if current_tokens + para_tokens > max_tokens:
                                if current_chunk:
                                    chunk_parts.append(current_chunk)
                                    current_chunk = para
                                    current_tokens = para_tokens
                                else:
                                    # Single paragraph exceeds limit
                                    chunk_parts.append(para)
                            else:
                                current_chunk += "\n\n" + para if current_chunk else para
                                current_tokens += para_tokens
                        
                        # Add the last chunk if not empty
                        if current_chunk:
                            chunk_parts.append(current_chunk)
                        
                        # Create a chunk for each part
                        for j, part in enumerate(chunk_parts):
                            part_tokens = count_tokens(part, model)
                            chunk = SmartChunk(
                                chunk_id=len(chunks) + 1,  # 1-indexed
                                policy_name=policy.name,
                                content=part,
                                pages=policy.page_range.page_list,
                                tokens=part_tokens,
                                complete_policy=False,
                                chunk_part=j + 1,
                                total_chunks=len(chunk_parts),
                                is_vision_extracted=content_result.get("is_vision_extracted", False)
                            )
                            chunk_dict = chunk.dict()
                            
                            # Add tables and table flag if present - but only to the first chunk
                            if j == 0:
                                if content_result.get("contains_tables", False):
                                    chunk_dict["contains_tables"] = True
                                if "tables" in content_result and content_result["tables"]:
                                    chunk_dict["tables"] = content_result["tables"]
                            
                            # Add Arabic content if available
                            if is_arabic and "arabic_content" in content_result:
                                chunk_dict["arabic_content"] = content_result["arabic_content"]
                            
                            chunks.append(chunk_dict)
                    else:
                        # We can split by pages
                        current_chunk = ""
                        current_tokens = 0
                        current_pages = []
                        chunk_parts = []
                        chunk_page_lists = []
                        
                        for j, (page_text, page_num) in enumerate(zip(page_contents, page_splits)):
                            page_text = f"[PAGE {page_num}]\n{page_text}"
                            page_tokens = count_tokens(page_text, model)
                            page_num = int(page_num)
                            
                            if current_tokens + page_tokens > max_tokens:
                                if current_chunk:
                                    chunk_parts.append(current_chunk)
                                    chunk_page_lists.append(current_pages)
                                    current_chunk = page_text
                                    current_tokens = page_tokens
                                    current_pages = [page_num - 1]  # Store 0-indexed page
                                else:
                                    # Single page exceeds limit
                                    chunk_parts.append(page_text)
                                    chunk_page_lists.append([page_num - 1])
                            else:
                                current_chunk += "\n\n" + page_text if current_chunk else page_text
                                current_tokens += page_tokens
                                current_pages.append(page_num - 1)  # Store 0-indexed page
                        
                        # Add the last chunk if not empty
                        if current_chunk:
                            chunk_parts.append(current_chunk)
                            chunk_page_lists.append(current_pages)
                        
                        # Create a chunk for each part
                        for j, (part, pages) in enumerate(zip(chunk_parts, chunk_page_lists)):
                            part_tokens = count_tokens(part, model)
                            chunk = SmartChunk(
                                chunk_id=len(chunks) + 1,  # 1-indexed
                                policy_name=policy.name,
                                content=part,
                                pages=pages,
                                tokens=part_tokens,
                                complete_policy=False,
                                chunk_part=j + 1,
                                total_chunks=len(chunk_parts),
                                is_vision_extracted=content_result.get("is_vision_extracted", False)
                            )
                            chunk_dict = chunk.dict()
                            
                            # Add tables and table flag if present
                            # For page-based splitting, we need to check which tables belong to which pages
                            if content_result.get("contains_tables", False):
                                chunk_dict["contains_tables"] = True
                            
                            if "tables" in content_result and content_result["tables"]:
                                # Only include tables that are on pages in this chunk
                                chunk_tables = [
                                    table for table in content_result["tables"]
                                    if table.get("page") - 1 in pages  # Convert 1-indexed page to 0-indexed
                                ]
                                if chunk_tables:
                                    chunk_dict["tables"] = chunk_tables
                            
                            # Add Arabic content if available - should be split by pages as well
                            if is_arabic and "arabic_content" in content_result:
                                # This is a simplification - in a real implementation we should 
                                # split the Arabic content by pages too
                                chunk_dict["arabic_content"] = content_result["arabic_content"]
                            
                            chunks.append(chunk_dict)
        
        logger.info(f"Created {len(chunks)} smart chunks from {len(doc_structure.policies)} policies")
        return chunks
        
    except Exception as e:
        logger.error(f"Error creating smart chunks: {e}")
        return []


def save_policies_to_json(policies: List[Dict[str, Any]], output_dir: str, doc_path: str, analysis_result: Dict = None) -> None:
    """
    Save policies to disk in JSON format with enhanced metadata
    """
    # Get filename without extension
    base_filename = os.path.splitext(os.path.basename(doc_path))[0]
    # Clean the base filename to remove special characters
    base_filename = clean_filename(base_filename)
    
    # Extract parent folder name (e.g., "DMT") from the input path
    parent_folder = os.path.basename(os.path.dirname(doc_path)) 
    
    # Create document-specific directory including parent folder name
    if parent_folder and parent_folder != ".":
        doc_output_dir = os.path.join(output_dir, parent_folder, base_filename)
    else:
        doc_output_dir = os.path.join(output_dir, base_filename)
    
    # Clean up existing directory
    if os.path.exists(doc_output_dir):
        logger.info(f"Removing existing output directory: {doc_output_dir}")
        shutil.rmtree(doc_output_dir)
    
    os.makedirs(doc_output_dir, exist_ok=True)
    
    # Create policies directory
    policies_dir = os.path.join(doc_output_dir, "policies")
    os.makedirs(policies_dir, exist_ok=True)
    
    # Create document-level metadata
    doc_metadata = {
        "document": {
            "filename": os.path.basename(doc_path),
            "path": doc_path,
            "parent_folder": parent_folder,  # Add parent folder to metadata
            "processing_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "language": analysis_result.get("language_info", {}).get("detected_language", "en"),
            "is_arabic": analysis_result.get("language_info", {}).get("is_arabic", False)
        },
        "policies": {
            "total_count": len(policies),
            "policy_count": sum(1 for p in policies if p.get("is_policy", True)),
            "non_policy_count": sum(1 for p in policies if not p.get("is_policy", True)),
            "vision_extracted_count": sum(1 for p in policies if p.get("metadata", {}).get("is_vision_extracted", False)),
            "summary": [
                {
                    "policy_id": policy.get("policy_id", i+1),  # Fallback to index if no ID
                    "policy_name": policy.get("policy_name", "Untitled Policy"),
                    "policy_name_arabic": policy.get("policy_name_arabic", ""),
                    "is_policy": policy.get("is_policy", True),
                    "page_range": policy.get("page_range", []),
                    "policy_type": policy.get("metadata", {}).get("policy_type", "unknown"),
                    "is_vision_extracted": policy.get("metadata", {}).get("is_vision_extracted", False)
                }
                for i, policy in enumerate(policies)
            ]
        }
    }
    
    # Save the vision analysis
    if analysis_result:
        doc_metadata["vision_analysis"] = analysis_result
        vision_file = os.path.join(doc_output_dir, "vision_analysis.json")
        with open(vision_file, "w") as f:
            json.dump(analysis_result, f, indent=2)
        logger.info(f"Saved vision analysis to {vision_file}")
    
    # Save document metadata
    with open(os.path.join(doc_output_dir, "document_metadata.json"), "w") as f:
        json.dump(doc_metadata, f, indent=2)
    
    # Save individual policy files
    for i, policy in enumerate(policies):
        policy_id = policy.get("policy_id", i+1)
        # Use English policy name for filenames
        policy_name = policy.get("policy_name", "Untitled Policy")
        
        # Calculate prefix length to ensure total length is 50 chars
        prefix = f"policy_{policy_id:03d}_"
        suffix = ".json"
        max_name_length = 50 - len(prefix) - len(suffix)
        
        # Limit policy name to remaining available characters
        truncated_name = clean_filename(policy_name[:max_name_length])
        
        # Create filename that's guaranteed to be at most 50 characters
        policy_filename = f"{prefix}{truncated_name}{suffix}"
        
        # Verify filename length and log if truncated significantly
        if len(truncated_name) < len(clean_filename(policy_name)):
            logger.info(f"Truncated policy filename from {len(clean_filename(policy_name))} to {len(truncated_name)} chars")
        
        with open(os.path.join(policies_dir, policy_filename), "w") as f:
            json.dump(policy, f, indent=2)
    
    logger.info(f"Saved {len(policies)} policies as JSON to {doc_output_dir}")


def policies_to_dataframe(policies: List[Dict[str, Any]], doc_path: str) -> pd.DataFrame:
    """
    Convert policy data to a pandas DataFrame for export to Excel
    
    Args:
        policies: List of policy dictionaries
        doc_path: Path to the source document
        
    Returns:
        DataFrame with policy data
    """
    # Extract data for each policy
    policy_data = []
    
    # Get relative path from current directory
    rel_path = os.path.relpath(doc_path)
    filename = os.path.basename(doc_path)
    
    for policy in policies:
        # Extract key metadata
        policy_id = policy.get("policy_id", 0)
        policy_name = policy.get("policy_name", "Untitled Policy")
        is_policy = policy.get("is_policy", True)
        doc_page = policy.get("document_page", 0)
        pdf_page = policy.get("pdf_page", 0)
        page_range = policy.get("page_range", [])
        page_range_str = ", ".join(map(str, page_range)) if page_range else ""
        policy_type = policy.get("metadata", {}).get("policy_type", "unknown")
        contains_tables = policy.get("metadata", {}).get("contains_tables", False)
        year = policy.get("metadata", {}).get("year", None)
        token_count = policy.get("metadata", {}).get("tokens", 0)
        
        # First ~200 characters of content as preview
        content_preview = policy.get("content", "")[:200] + "..." if policy.get("content") else ""
        
        # Add to data list
        policy_data.append({
            "document_path": rel_path,
            "filename": filename,
            "policy_id": policy_id,
            "policy_name": policy_name,
            "is_policy": is_policy,
            "policy_type": policy_type,
            "document_page": doc_page,
            "pdf_page": pdf_page,
            "page_range": page_range_str,
            "contains_tables": contains_tables,
            "year": year,
            "token_count": token_count,
            "content_preview": content_preview
        })
    
    # Create DataFrame
    df = pd.DataFrame(policy_data)
    
    # Reorder columns for better readability
    column_order = [
        "document_path", "filename", "policy_id", "policy_name", "is_policy", 
        "policy_type", "year", "document_page", "pdf_page", "page_range",
        "contains_tables", "token_count", "content_preview"
    ]
    
    # Ensure all columns exist (some might be missing in certain scenarios)
    existing_columns = [col for col in column_order if col in df.columns]
    df = df[existing_columns]
    
    return df


def save_policies_to_excel(policies: List[Dict[str, Any]], output_dir: str, doc_path: str) -> str:
    """
    Save policies to an Excel file for easy reviewing.
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    # Get document name without extension
    doc_name = os.path.splitext(os.path.basename(doc_path))[0]
    # Clean the document name to remove special characters
    doc_name = clean_filename(doc_name)
    
    # Extract parent folder name (e.g., "DMT") from the input path
    parent_folder = os.path.basename(os.path.dirname(doc_path))
    
    # Create subdirectory for this document, including parent folder
    if parent_folder and parent_folder != ".":
        doc_dir = os.path.join(output_dir, parent_folder, doc_name)
    else:
        doc_dir = os.path.join(output_dir, doc_name)
    
    os.makedirs(doc_dir, exist_ok=True)
    
    # Create Excel file path
    excel_path = os.path.join(doc_dir, f"{doc_name}_policies.xlsx")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default worksheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create summary worksheet
    summary_sheet = wb.create_sheet("Summary")
    
    # Add headers
    summary_headers = [
        "Policy ID", "Policy Name", "Arabic Name", "Is Policy", "Document Page", "PDF Page",
        "Page Range", "Language", "Is Arabic", "Year", "Policy Type", "Contains Tables", 
        "Contains Images", "Tokens", "Relative Path"
    ]
    summary_sheet.append(summary_headers)
    
    # Set header style
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
    
    # Add policy data
    for idx, policy in enumerate(policies):
        metadata = policy.get("metadata", {})
        
        # Get English policy name for filename (already translated if Arabic)
        policy_name = policy.get("policy_name", "")
        
        # Calculate relative path to individual policy file - use English name for filename
        rel_path = f"./policies/policy_{policy.get('policy_id', 0):03d}_{clean_filename(policy_name[:50])}.json"
        
        # Format page range as a string
        page_range = ', '.join(map(str, policy.get('page_range', [])))
        
        summary_row = [
            policy.get("policy_id"),
            policy_name,
            policy.get("policy_name_arabic", ""),  # Include Arabic name if available
            policy.get("is_policy", True),
            policy.get("document_page"),
            policy.get("pdf_page"),
            page_range,
            metadata.get("language", ""),
            metadata.get("is_arabic", False),
            metadata.get("year"),
            metadata.get("policy_type", "unknown"),
            metadata.get("contains_tables", False),
            metadata.get("contains_images", False),
            metadata.get("tokens", 0),
            rel_path
        ]
        summary_sheet.append(summary_row)
        
        # Apply RTL alignment for Arabic content
        row_idx = idx + 2  # +2 because Excel is 1-indexed and we have a header row
        if policy.get("policy_name_arabic"):
            # Set RTL alignment for Arabic name cell
            arabic_name_cell = summary_sheet.cell(row=row_idx, column=3)  # Column 3 is "Arabic Name"
            try:
                # Try with reading_order parameter for newer openpyxl versions
                arabic_name_cell.alignment = Alignment(horizontal='right', vertical='center', reading_order=2)  # reading_order=2 for RTL
            except TypeError:
                # Fallback for older openpyxl versions
                arabic_name_cell.alignment = Alignment(horizontal='right', vertical='center')
                logger.info("Using simplified RTL alignment (older openpyxl version detected)")
    
    # Adjust column widths
    for col in summary_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) if max_length < 80 else 80
        summary_sheet.column_dimensions[column].width = adjusted_width
    
    # Create English content sheet
    content_sheet = wb.create_sheet("Content")
    
    # Add headers for content sheet
    content_headers = [
        "Policy ID", "Policy Name", "Is Policy", "Content",
    ]
    content_sheet.append(content_headers)
    
    # Set header style
    for cell in content_sheet[1]:
        cell.font = Font(bold=True)
    
    # Add policy content (always the main content - translated for Arabic documents)
    for policy in policies:
        content_row = [
            policy.get("policy_id"),
            policy.get("policy_name"),
            policy.get("is_policy", True),
            policy.get("content", "")
        ]
        content_sheet.append(content_row)
    
    # Adjust column widths
    for col in content_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if column != 'D':  # Skip the content column
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
        if column != 'D':  # Skip the content column
            adjusted_width = (max_length + 2) if max_length < 80 else 80
            content_sheet.column_dimensions[column].width = adjusted_width
    
    # Set content column width
    content_sheet.column_dimensions['D'].width = 120
    
    # Create Arabic content sheet if needed
    has_arabic = any(policy.get("policy_name_arabic") or policy.get("arabic_content") for policy in policies)
    if has_arabic:
        arabic_sheet = wb.create_sheet("Arabic Content")
        
        # Add headers for Arabic sheet
        arabic_headers = [
            "Policy ID", "English Name", "Arabic Name", "Arabic Content",
        ]
        arabic_sheet.append(arabic_headers)
        
        # Set header style
        for cell in arabic_sheet[1]:
            cell.font = Font(bold=True)
        
        # Add Arabic policy content
        for idx, policy in enumerate(policies):
            if policy.get("policy_name_arabic") or policy.get("arabic_content"):
                arabic_row = [
                    policy.get("policy_id"),
                    policy.get("policy_name"),
                    policy.get("policy_name_arabic", ""),
                    policy.get("arabic_content", "")
                ]
                arabic_sheet.append(arabic_row)
                
                try:
                    # Apply RTL alignment for Arabic content
                    row_idx = idx + 2  # +2 because Excel is 1-indexed and we have a header row
                    # Set RTL alignment for Arabic name cell
                    arabic_name_cell = arabic_sheet.cell(row=row_idx, column=3)  # Column 3 is "Arabic Name"
                    arabic_name_cell.alignment = Alignment(horizontal='right', vertical='center', reading_order=2)  # reading_order=2 for RTL
                    
                    # Set RTL alignment for Arabic content cell
                    arabic_content_cell = arabic_sheet.cell(row=row_idx, column=4)  # Column 4 is "Arabic Content"
                    arabic_content_cell.alignment = Alignment(horizontal='right', vertical='top', reading_order=2, wrap_text=True)  # reading_order=2 for RTL
                except TypeError:
                    # Fallback for older openpyxl versions
                    arabic_name_cell = arabic_sheet.cell(row=row_idx, column=3)
                    arabic_name_cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    arabic_content_cell = arabic_sheet.cell(row=row_idx, column=4)
                    arabic_content_cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
        
        # Adjust column widths
        for col in arabic_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if column != 'D':  # Skip the content column
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            if column != 'D':  # Skip the content column
                adjusted_width = (max_length + 2) if max_length < 80 else 80
                arabic_sheet.column_dimensions[column].width = adjusted_width
        
        # Set content column width
        arabic_sheet.column_dimensions['D'].width = 120
    
    # Save workbook
    wb.save(excel_path)
    logger.info(f"Saved policy information to Excel file: {excel_path}")
    
    return excel_path


def append_to_master_excel(source_excel_path, master_excel_path):
    """Append policy data from a document Excel file to the master Excel file."""
    import pandas as pd
    from openpyxl import load_workbook
    
    # Try to read the source Excel file
    try:
        source_wb = load_workbook(source_excel_path)
        # Use modern syntax to avoid deprecation warning
        source_sheet = source_wb["Summary"]
    except Exception as e:
        logger.error(f"Failed to read source Excel file: {source_excel_path} - Error: {e}")
        return
    
    # Convert source sheet to DataFrame
    source_data = []
    
    # Get headers from first row
    headers = [cell.value for cell in source_sheet[1]]
    
    # Get data from remaining rows
    for row in source_sheet.iter_rows(min_row=2):
        row_data = [cell.value for cell in row]
        source_data.append(dict(zip(headers, row_data)))
    
    source_df = pd.DataFrame(source_data)
    
    # Extract document information from the path
    source_path = os.path.dirname(source_excel_path)
    doc_name = os.path.basename(os.path.dirname(source_path))  # Get parent folder
    collection_name = os.path.basename(source_path)  # Get document folder name
    
    # If we're using the new structure with parent folders
    if doc_name == "output" or doc_name == "Output":
        # Old structure - use just the collection name
        doc_name = collection_name
    else:
        # New structure - use parent/collection format
        doc_name = f"{doc_name}/{collection_name}"
    
    # Add document and collection info to the DataFrame
    source_df["Document"] = doc_name
    source_df["Collection"] = collection_name
    
    # Create a relative path to the policy JSON files
    source_df["Relative Path"] = source_df.apply(
        lambda row: f"./{doc_name}/policies/policy_{row['Policy ID']:03d}_{clean_filename(row['Policy Name'][:40])}.json" 
        if pd.notna(row.get('Policy ID')) else "", 
        axis=1
    )
    
    # List of tracking columns to preserve
    tracking_columns = ["processed"]
    
    try:
        # Check if master Excel file exists and read it
        if os.path.exists(master_excel_path):
            # Read existing master Excel file
            master_df = pd.read_excel(master_excel_path)
            
            # Store existing tracking information
            tracking_info = {}
            
            # Create a dictionary to store tracking data based on unique identifiers
            if not master_df.empty:
                for index, row in master_df.iterrows():
                    # Create a unique identifier for each policy (Document + Policy ID)
                    if pd.notna(row.get('Document')) and pd.notna(row.get('Policy ID')):
                        unique_id = f"{row['Document']}_{row['Policy ID']}"
                        
                        # Store tracking data for this policy (only processed column)
                        policy_tracking = {}
                        for col in tracking_columns:
                            if col in master_df.columns and pd.notna(row.get(col)):
                                policy_tracking[col] = row[col]
                        
                        if policy_tracking:  # Only add if we have some tracking data
                            tracking_info[unique_id] = policy_tracking
            
            # Add tracking columns to source_df but leave them explicitly blank (None/NaN)
            for col in tracking_columns:
                if col not in source_df.columns:
                    source_df[col] = None
            
            # Create a list of existing unique IDs for comparison
            existing_ids = []
            if not master_df.empty:
                for index, row in master_df.iterrows():
                    if pd.notna(row.get('Document')) and pd.notna(row.get('Policy ID')):
                        existing_ids.append(f"{row['Document']}_{row['Policy ID']}")
            
            # Flag new entries in source_df
            for index, row in source_df.iterrows():
                if pd.notna(row.get('Document')) and pd.notna(row.get('Policy ID')):
                    unique_id = f"{row['Document']}_{row['Policy ID']}"
                    
                    # If this is an existing entry with tracking data, restore it
                    if unique_id in tracking_info:
                        for col, value in tracking_info[unique_id].items():
                            source_df.at[index, col] = value
            
            # Append the new data
            combined_df = pd.concat([master_df, source_df], ignore_index=True)
            
            # Remove duplicates, but be careful about the 'processed' column:
            # For duplicates, keep the row that has tracking data (processed value) if it exists
            def keep_row_with_tracking(group):
                # If any row has a non-null processed value, keep that one
                if any(pd.notna(group['processed'])):
                    return group[pd.notna(group['processed'])].iloc[0]
                # Otherwise keep the latest entry (which would be from source_df)
                return group.iloc[-1]
            
            # Group by the identifying columns and apply our custom function
            combined_df = combined_df.drop_duplicates(
                subset=["Document", "Policy ID", "Policy Name"], 
                keep="first"  # Keep first occurrence (from master_df) to preserve tracking data
            )
            # Write the combined DataFrame to the master Excel file
            combined_df.to_excel(master_excel_path, index=False)
            logger.info(f"Updated master Excel file with {len(source_df)} policies: {master_excel_path}")
            
        else:
            # If master Excel file doesn't exist, make sure we have the essential tracking columns
            # but leave them explicitly blank
            for col in tracking_columns:
                if col not in source_df.columns:
                    source_df[col] = None
                    
            # Write source DataFrame directly to master Excel file
            source_df.to_excel(master_excel_path, index=False)
            logger.info(f"Created new master Excel file with {len(source_df)} policies: {master_excel_path}")
            
    except Exception as e:
        logger.error(f"Failed to update master Excel file: {master_excel_path} - Error: {e}")
        logger.exception(e)


def extract_tables_with_vision(client, doc_path, page_num, table_regions=None, metadata=None):
    """
    Extract tables from a PDF page using the vision model.
    Uses the detect_tables_in_pdf function from utils.py if table_regions not provided.
    
    Args:
        client: Azure OpenAI client
        doc_path: Path to the PDF document
        page_num: Page number to extract tables from (0-indexed)
        table_regions: Optional list of table regions [(x0, y0, x1, y1), ...] 
        metadata: Additional metadata about the policy
        
    Returns:
        List of extracted tables as markdown strings
    """
    try:
        # First check if the page is valid
        try:
            with fitz.open(doc_path) as test_doc:
                total_pages = len(test_doc)
                if page_num < 0 or page_num >= total_pages:
                    logger.warning(f"Page {page_num+1} is out of bounds (document has {total_pages} pages)")
                    return []
        except Exception as e:
            logger.error(f"Error opening document to check page bounds: {e}")
            return []
            
        # Set up a sync client if needed
        sync_client = setup_azure_openai_client(async_client=False)
        
        # Extract is_arabic from metadata
        is_arabic = metadata.get('is_arabic', False) if metadata else False
        
        # If no table regions provided, detect them
        if not table_regions:
            try:
                # Import detect_tables_in_pdf from utils (already imported at the top)
                tables_by_page = detect_tables_in_pdf(doc_path, page_numbers=[page_num])
                if page_num not in tables_by_page or not tables_by_page[page_num]:
                    logger.info(f"No tables detected on page {page_num+1}")
                    return []
                
                # Filter high-confidence tables
                high_conf_tables = [
                    table for table in tables_by_page[page_num] 
                    if is_valid_table(table)
                ]
                
                if not high_conf_tables:
                    logger.info(f"No high-confidence tables on page {page_num+1}")
                    return []
                    
                table_regions = [table["bbox"] for table in high_conf_tables]
                logger.info(f"Using {len(table_regions)} detected table regions on page {page_num+1}")
            except Exception as e:
                logger.warning(f"Error detecting tables: {e}. Will try to extract without table detection.")
                return []
        
        # If we still have no table regions, return empty result
        if not table_regions:
            return []
        
        # Open the document and get the page
        try:
            doc = fitz.open(doc_path)
            page = doc[page_num]
            
            # Convert page to image
            img = page_to_image(page)
            width, height = img.size
            
            # Process each table region
            extracted_tables = []
            
            def extract_one(region):
                try:
                    x0, y0, x1, y1 = region
                    x0 = max(0, min(x0, width)); y0 = max(0, min(y0, height))
                    x1 = max(0, min(x1, width)); y1 = max(0, min(y1, height))
                    if x1 <= x0 or y1 <= y0 or (x1 - x0) < 20 or (y1 - y0) < 20:
                        return None
                    table_img = img.crop((x0, y0, x1, y1))
                    buffer = BytesIO(); table_img.save(buffer, format="JPEG", quality=70, optimize=True)
                    base64_image = base64.b64encode(buffer.getvalue()).decode('utf-8')
                    prompt = "Extract this table as Markdown format. Maintain the same columns and rows. ONLY return the markdown table, nothing else."
                    if is_arabic:
                        prompt += " This table may contain Arabic text. Please preserve the Arabic content exactly as shown."
                    response = sync_client.chat.completions.create(
                        model="gpt-5",
                        messages=[
                            {"role": "system", "content": "You are a table extraction specialist. Your task is to convert images of tables into clean, structured markdown format."},
                            {"role": "user", "content": [
                                {"type": "text", "text": prompt},
                                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}", "detail": "high"}}
                            ]}
                        ],
                        temperature=0.0,
                        seed=42
                    )
                    table_markdown = response.choices[0].message.content.strip()
                    if '|' in table_markdown and '-' in table_markdown:
                        return table_markdown
                    return None
                except Exception as e:
                    return None

            with ThreadPoolExecutor(max_workers=TABLE_VISION_CONCURRENCY) as ex:
                futs = [ex.submit(extract_one, region) for region in table_regions]
                for i, fut in enumerate(as_completed(futs), start=1):
                    md = fut.result()
                    if md:
                        extracted_tables.append(md)
                        logger.info(f"Successfully extracted a table from page {page_num+1} ({len(md)} chars)")
            
            # Clean up
            doc.close()
            return extracted_tables
            
        except Exception as e:
            logger.error(f"Error in table extraction process: {e}")
            if 'doc' in locals():
                doc.close()
            return []
            
    except Exception as e:
        logger.error(f"Error in vision-based table extraction: {e}")
        return []


def check_file_already_processed(file_path, master_excel_path):
    """
    Check if a file has already been processed by examining the master Excel file.
    
    Args:
        file_path: Path to the PDF file to check
        master_excel_path: Path to the master Excel file (all_policies.xlsx)
        
    Returns:
        bool: True if file has already been processed, False otherwise
    """
    if not os.path.exists(master_excel_path):
        logger.info(f"Master Excel file not found at {master_excel_path}. Will process all files.")
        return False
        
    try:
        # Get filename and parent folder for matching
        filename = os.path.basename(file_path)
        parent_folder = os.path.basename(os.path.dirname(file_path))
        
        # Read the master Excel file
        master_df = pd.read_excel(master_excel_path)
        
        # Log columns for debugging
        logger.debug(f"Columns in all_policies.xlsx: {list(master_df.columns)}")
        
        # Check for the filename column - it might be 'filename' or 'Filename' or other variants
        filename_col = None
        possible_filename_cols = ['filename', 'Filename', 'FileName', 'file_name', 'File Name', 'file']
        for col in possible_filename_cols:
            if col in master_df.columns:
                filename_col = col
                break
                
        # If we can't find a filename column, try to use a different approach
        if filename_col is None:
            logger.warning(f"Could not find filename column in {master_excel_path}. Looking for filename in document path.")
            
            # Try to match based on the filename in the document path
            if 'document_path' in master_df.columns or 'Document Path' in master_df.columns:
                doc_path_col = 'document_path' if 'document_path' in master_df.columns else 'Document Path'
                
                # Check if any document path contains this filename
                matches = master_df[master_df[doc_path_col].str.contains(filename, na=False)]
                
                if not matches.empty:
                    logger.info(f"Found file {filename} in document paths. Skipping.")
                    return True
            elif 'Document' in master_df.columns:
                # Try the Document column which might contain the filename or folder
                matches = master_df[master_df['Document'].str.contains(filename.replace('.pdf', ''), na=False)]
                
                if not matches.empty:
                    logger.info(f"Found file {filename} in Document column. Skipping.")
                    return True
            else:
                logger.warning(f"Could not find suitable columns to check if file was processed. Will process file.")
                return False
        else:
            # We found a filename column, use it to check
            file_entries = master_df[master_df[filename_col] == filename]
            
            if file_entries.empty:
                logger.info(f"File {filename} not found in master Excel. Will process it.")
                return False
                
            # Find the Document or similar column for checking folder
            document_col = None
            possible_document_cols = ['Document', 'document', 'Collection', 'collection', 'folder']
            for col in possible_document_cols:
                if col in master_df.columns:
                    document_col = col
                    break
            
            # If parent folder is relevant and we have a document column, check if it matches
            if parent_folder and parent_folder != "." and document_col is not None:
                # Check entries with matching parent folder
                folder_matches = file_entries[file_entries[document_col].str.contains(parent_folder, na=False)]
                
                if folder_matches.empty:
                    logger.info(f"File {filename} found in master Excel but with different folder. Will process it.")
                    return False
                    
                logger.info(f"File {filename} already processed in folder {parent_folder}. Skipping.")
                return True
            else:
                # If we found entries for this filename without checking folder
                logger.info(f"File {filename} already processed. Skipping.")
                return True
            
        # If we got this far without returning, process the file
        return False
            
    except Exception as e:
        logger.warning(f"Error checking if file was processed: {e}. Will process file.")
        return False


async def main():
    """Main function for CLI usage"""
    parser = argparse.ArgumentParser(description='Smart document chunking for policy extraction')
    parser.add_argument('--input', '-i', required=True, help='Path to PDF document or directory containing PDF documents')
    parser.add_argument('--output', '-o', default='output', help='Output directory for chunks')
    parser.add_argument('--max-tokens', '-m', type=int, default=100000, help='Maximum tokens per chunk')
    parser.add_argument('--model', type=str, default='gpt-5', help='Model for token counting')
    parser.add_argument('--toc-pages', type=int, default=5, help='Number of pages to analyze for table of contents')
    parser.add_argument('--no-excel', action='store_true', help='Skip Excel file generation')
    parser.add_argument('--no-translation', action='store_true', help='Skip translation for Arabic documents')
    parser.add_argument('--no-tables', action='store_true', help='Skip table extraction')
    parser.add_argument('--force', action='store_true', help='Force processing even if file was already processed')
    parser.add_argument('--limit', '-l', type=int, default=None, help='Limit the number of files to process from a directory')
    parser.add_argument('--skip-vision-for-tables',default=True, action='store_true', help='Skip using vision model when tables are detected')
    parser.add_argument('--debug', action='store_true', help='Enable debug logging')
    parser.add_argument('--quiet', action='store_true', help='Reduce logging output')
    parser.add_argument('--progress', action='store_true', help='Show detailed progress tracking', default=True)
    
    args = parser.parse_args()
    
    # Configure logging level based on arguments
    if args.debug:
        logger.setLevel(logging.DEBUG)
    elif args.quiet:
        logger.setLevel(logging.WARNING)
    else:
        logger.setLevel(logging.INFO)
    
    # Path to the master Excel file
    master_excel_path = os.path.join(args.output, "all_policies.xlsx")

    print(f"Master Excel path: {master_excel_path}")
    
    # Check if input is a directory or a file
    if os.path.isdir(args.input):
        # Process all PDF files in the directory
        pdf_files = [os.path.join(args.input, f) for f in os.listdir(args.input) 
                    if f.lower().endswith('.pdf') and os.path.isfile(os.path.join(args.input, f))]
        
        if not pdf_files:
            logger.error(f"No PDF files found in directory: {args.input}")
            return
        
        # Initialize progress tracking
        progress_tracker = None
        if args.progress:
            try:
                from progress_tracker import ProgressTracker
                progress_tracker = ProgressTracker(len(pdf_files), "Smart Chunking")
            except ImportError:
                logger.warning("Progress tracker not available")
        
        logger.info(f"Found {len(pdf_files)} PDF files in directory: {args.input}")
        
        # Process each PDF file
        processed_count = 0
        skipped_count = 0
        
        for pdf_file in pdf_files:
            # Check if this file has already been processed
            if not args.force and check_file_already_processed(pdf_file, master_excel_path):
                if progress_tracker:
                    progress_tracker.update(os.path.basename(pdf_file), "Already processed", True)
                logger.info(f"Skipping already processed file: {pdf_file}")
                skipped_count += 1
                continue
                
            # Apply limit to files actually being processed
            if args.limit is not None and processed_count >= args.limit:
                logger.info(f"Reached limit of {args.limit} processed files. Stopping.")
                break
            
            # Update progress
            if progress_tracker:
                progress_tracker.update(os.path.basename(pdf_file), "Analyzing document structure")
                
            logger.info(f"Processing file {processed_count + 1}/{min(len(pdf_files), args.limit if args.limit is not None else len(pdf_files))}: {pdf_file}")
            
            try:
                await process_single_file(pdf_file, args)
                if progress_tracker:
                    progress_tracker.update("", "Completed", True)
                processed_count += 1
            except Exception as e:
                if progress_tracker:
                    progress_tracker.update("", "Failed", False)
                logger.error(f"Error processing {pdf_file}: {e}")
                processed_count += 1  # Still count as processed
            
        # Complete progress tracking
        if progress_tracker:
            progress_tracker.complete(f"Processed {processed_count} files, skipped {skipped_count} files")
        
        logger.info(f"Completed processing: {processed_count} files processed, {skipped_count} files skipped")
    else:
        # Process a single file
        if not os.path.exists(args.input):
            logger.error(f"Input file not found: {args.input}")
            return
        
        if not args.input.lower().endswith('.pdf'):
            logger.error(f"Input file is not a PDF: {args.input}")
            return
        
        # For single file processing, check if already processed unless force flag is set
        if not args.force and check_file_already_processed(args.input, master_excel_path):
            logger.info(f"File already processed. Use --force to process again.")
            return
            
        await process_single_file(args.input, args)


async def process_single_file(doc_path, args):
    """Process a single PDF file"""
    logger.info(f"Processing document: {doc_path}")
    
    # Set up OpenAI client
    client = setup_azure_openai_client()
    
    # STEP 1: Analyze document structure with vision using initial pages
    logger.info(f"Analyzing document structure...")
    analysis_result = await analyze_initial_pages_with_vision(
        client, doc_path, max_pages=args.toc_pages
    )
    
    # Get language information
    language_info = analysis_result.get("language_info", {})
    is_arabic = language_info.get("is_arabic", False)
    logger.info(f"Document language: {'Arabic' if is_arabic else language_info.get('detected_language', 'English')}")
    
    # Print summary of policies found
    if "policies" in analysis_result and analysis_result["policies"]:
        page_offset = analysis_result.get("page_offset", 1)
        policies = analysis_result["policies"]
        logger.info(f"Found {len(policies)} policies with page offset: {page_offset}")
    else:
        logger.warning("No policies found in vision analysis")
    
    # STEP 2: Create smart chunks from document structure
    logger.info("Creating smart chunks from document structure...")
    chunks = await create_smart_chunks(
        doc_path, 
        args.max_tokens, 
        args.model, 
        analysis_result, 
        client,
        args.skip_vision_for_tables
    )
    
    # Early return if no chunks were created
    if not chunks:
        logger.error("No chunks could be created. Check if the document contains valid content.")
        return
    
    # STEP 3: Process policies from vision analysis
    full_policies = []
    if "policies" in analysis_result and analysis_result["policies"]:
        # Create a mapping of policy name to chunks
        policy_chunks = {}
        for chunk in chunks:
            policy_name = chunk["policy_name"]
            if policy_name not in policy_chunks:
                policy_chunks[policy_name] = []
            policy_chunks[policy_name].append(chunk)
        
        # Process each policy from the vision analysis
        page_offset = analysis_result.get("page_offset", 1)
        
        for i, policy_data in enumerate(analysis_result["policies"]):
            policy_name = policy_data.get("policy_name", "")
            policy_name_arabic = policy_data.get("policy_name_arabic", "")
            
            # Create the policy object with content from chunks
            doc_page = int(policy_data.get("page_number", 0))
            pdf_page = (doc_page + page_offset) - 1  # Add offset and convert to 0-indexed
            
            # Find chunks for this policy
            policy_content_chunks = policy_chunks.get(policy_name, [])
            if not policy_content_chunks:
                logger.warning(f"No chunks found for policy: {policy_name}")
                continue
            
            # Gather pages and content from chunks
            pages = []
            full_content = ""
            has_tables = False
            arabic_content = None
            tables = []
            
            for chunk in policy_content_chunks:
                pages.extend(chunk.get("pages", []))
                full_content += chunk.get("content", "") + "\n\n"
                
                # Check for tables in this chunk
                if chunk.get("contains_tables", False):
                    has_tables = True
                
                # Collect tables if available
                if "tables" in chunk:
                    tables.extend(chunk["tables"])
                
                # Get Arabic content if available
                if is_arabic and "arabic_content" in chunk:
                    if arabic_content is None:
                        arabic_content = ""
                    arabic_content += chunk.get("arabic_content", "") + "\n\n"
            
            # Extract year from policy name if possible
            year_match = re.search(r'(?:of|لسنة)\s*(\d{4})', policy_name if not policy_name_arabic else policy_name_arabic)
            year = int(year_match.group(1)) if year_match else None
            
            # Create policy object
            policy = {
                "policy_id": i + 1,  # 1-indexed for user friendliness
                "policy_name": policy_name,
                "is_policy": policy_data.get("is_policy", True),
                "document_page": doc_page,
                "pdf_page": pdf_page,  # 1-indexed for display
                "page_range": sorted(list(set(pages))) if pages else [pdf_page],
                "content": full_content.strip(),
                "metadata": {
                    "contains_tables": has_tables,
                    "contains_images": False,  # Not detecting images yet
                    "language": language_info.get("detected_language", "en"),
                    "is_arabic": is_arabic,
                    "year": year,
                    "policy_type": policy_data.get("policy_type", "unknown"),
                    "tokens": sum(chunk.get("tokens", 0) for chunk in policy_content_chunks),
                    "is_vision_extracted": any(chunk.get("is_vision_extracted", False) for chunk in policy_content_chunks)
                }
            }
            
            # Add tables if we have any
            if tables:
                policy["tables"] = tables
                policy["metadata"]["table_count"] = len(tables)
                logger.info(f"Added {len(tables)} tables to policy: {policy_name}")
            
            # Add Arabic policy name and content if available
            if is_arabic:
                if policy_name_arabic:
                    policy["policy_name_arabic"] = policy_name_arabic
                if arabic_content:
                    policy["arabic_content"] = arabic_content.strip()
            
            # If content is too short, try direct vision extraction
            if len(policy["content"].strip()) < 200:
                logger.info(f"Minimal content for policy: {policy_name}. Attempting direct vision extraction...")
                
                start_page = min(policy["page_range"]) if policy["page_range"] else pdf_page
                end_page = max(policy["page_range"]) if policy["page_range"] else pdf_page
                
                # Create metadata dict for Arabic content storage
                vision_metadata = {"is_arabic": is_arabic}
                
                vision_content = extract_policy_content_with_vision(
                    client, 
                    doc_path, 
                    start_page, 
                    end_page, 
                    vision_metadata
                )
                
                if vision_content and len(vision_content.strip()) > len(policy["content"].strip()):
                    policy["content"] = vision_content
                    policy["metadata"]["is_vision_extracted"] = True
                    
                    # For Arabic, add original content from vision metadata
                    if is_arabic and vision_metadata.get("full_arabic_content"):
                        policy["arabic_content"] = vision_metadata["full_arabic_content"]
                    
                    # Check if the vision extraction found tables
                    if vision_metadata.get("contains_tables", False):
                        policy["metadata"]["contains_tables"] = True
            
            full_policies.append(policy)
        
        # STEP 4: Save policies to disk
        if full_policies:
            logger.info(f"Saving {len(full_policies)} policies to disk")
            save_policies_to_json(full_policies, args.output, doc_path, analysis_result)
            
            # Generate Excel files if not disabled
            if not args.no_excel:
                excel_path = save_policies_to_excel(full_policies, args.output, doc_path)
                append_to_master_excel(excel_path, os.path.join(args.output, "all_policies.xlsx"))
            
            logger.info(f"Processing complete: {len(full_policies)} policies extracted")
        else:
            logger.warning("No policies could be processed. Check if document contains valid policies.")
    else:
        logger.error("No policies found in document analysis")


if __name__ == "__main__":
    asyncio.run(main())